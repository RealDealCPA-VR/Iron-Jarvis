"""Document tools (§19).

Four tools that let agents work with a user's real files:

* ``read_document``     — extract text from ANY local path (reading the user's
  real files is the point), absolute or workspace-relative.
* ``write_document``    — create a document WITHIN the session workspace only.
* ``extract_pdf``       — read_document specialised to PDFs.
* ``convert_document``  — read any supported source and re-write it as any
  supported target format (csv<->xlsx keep real rows/cells).

``document_tools()`` is a plain factory (no platform needed).
"""

from __future__ import annotations

import asyncio
import csv
from pathlib import Path
from typing import Any

from ..core.fs_policy import fs_read_ok
from ..tools.base import Reversibility, Tool, ToolContext, ToolResult, safe_path
from ..tools.undo import make_file_descriptor, revert_workspace_file, sha256_bytes
from .pdf_markdown import MARKITDOWN_SUFFIXES, document_to_markdown
from .readers import SUPPORTED_READ, extract_text
from .writers import SUPPORTED_WRITE, write_document

#: Cap on tool output to keep large documents from flooding the context window.
_MAX_OUTPUT = 16_000


def _resolve_read_path(raw: str, ctx: ToolContext) -> Path:
    """Absolute paths are used as-is; relative paths resolve under the workspace."""
    p = Path(raw)
    if p.is_absolute():
        return p
    return (Path(ctx.workspace) / raw)


def _truncate(text: str) -> tuple[str, bool]:
    if len(text) <= _MAX_OUTPUT:
        return text, False
    note = f"\n\n... [truncated to {_MAX_OUTPUT} of {len(text)} characters]"
    return text[:_MAX_OUTPUT] + note, True


class ListFolderTool(Tool):
    name = "list_folder"
    reversibility = Reversibility.READONLY  # a listing has no side effect
    description = (
        "List a REAL folder anywhere on this machine (e.g. the user's Downloads "
        "or Documents) — name, size, modified time per entry, biggest first. Use "
        "ABSOLUTE paths for the user's actual files (the session workspace is "
        "only a scratch area). Reads are policy-gated."
    )
    permission_key = "list_folder"
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Absolute folder path (or workspace-relative)"},
            "limit": {"type": "number", "description": "Max entries (default 200)"},
        },
        "required": ["path"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        folder = _resolve_read_path(str(args.get("path", "")), ctx)
        ok, reason = fs_read_ok(str(folder))
        if not ok:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if not folder.is_dir():
            return ToolResult(ok=False, error=f"not a folder: {folder}")
        limit = max(1, min(int(args.get("limit") or 200), 1000))
        entries: list[tuple[str, bool, int, float]] = []
        try:
            for p in folder.iterdir():
                try:
                    st = p.stat()
                    entries.append((p.name, p.is_dir(), st.st_size, st.st_mtime))
                except OSError:
                    continue
        except OSError as exc:
            return ToolResult(ok=False, error=f"could not list {folder}: {exc}")
        entries.sort(key=lambda e: e[2], reverse=True)  # biggest first
        shown = entries[:limit]
        from datetime import datetime

        lines = [
            f"{'DIR  ' if is_dir else ''}{name}  —  {size:,} bytes  —  "
            f"{datetime.fromtimestamp(mtime):%Y-%m-%d %H:%M}"
            for name, is_dir, size, mtime in shown
        ]
        header = f"{folder} — {len(entries)} entries" + (
            f" (showing {len(shown)})" if len(shown) < len(entries) else ""
        )
        body = header + ("\n" + "\n".join(lines) if lines else "\n(empty)")
        return ToolResult(ok=True, output=body, data={"path": str(folder), "total": len(entries)})


class ReadDocumentTool(Tool):
    name = "read_document"
    reversibility = Reversibility.READONLY  # extraction has no side effect
    returns_untrusted_content = True  # a read file can carry planted instructions
    description = (
        "Extract text from a document of any type — PDF, Word (.docx), Excel "
        "(.xlsx), PowerPoint (.pptx), CSV, or plain text/code. Scanned "
        "(image-only) PDFs are OCR-transcribed via the vision model when one "
        "is available. May target ANY local path (absolute, or relative to "
        "the workspace)."
    )

    def __init__(self, router_resolver: "Any | None" = None) -> None:
        #: () -> the platform's ModelRouter — powers the scanned-PDF OCR
        #: fallback. Optional so a bare factory (tests) still constructs.
        self._router_resolver = router_resolver
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "page_range": {
                "type": "string",
                "description": (
                    "Optional 1-based page (PDF) / slide (PPTX) slice, e.g. "
                    "'2', '1-3', '2-', '1,4-6'. Ignored by other formats."
                ),
            },
            "sheet": {
                "description": (
                    "Optional worksheet to read from an .xlsx — a sheet NAME or "
                    "0-based index. Ignored by other formats."
                ),
            },
        },
        "required": ["path"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        path = _resolve_read_path(args["path"], ctx)
        allowed, reason = fs_read_ok(path)
        if not allowed:
            return ToolResult(ok=False, error=reason)
        try:
            text = await asyncio.to_thread(  # CPU-bound parse off the loop
                extract_text,
                path,
                page_range=args.get("page_range"),
                sheet=args.get("sheet"),
            )
        except Exception as exc:  # reading real files must never crash the runtime
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        # SCANNED-PDF fallback: image-only pages have no text layer — recover
        # via vision OCR (never the mock; the note names the method) instead
        # of handing the model empty silence about a real document.
        note = ""
        from .ocr import looks_scanned_pdf, ocr_pdf

        if self._router_resolver is not None and looks_scanned_pdf(path, text):
            try:
                ocr_text, note = await ocr_pdf(path, self._router_resolver())
                if ocr_text:
                    text = ocr_text
            except Exception as exc:  # noqa: BLE001 — OCR failure ≠ read failure
                note = f"scanned PDF — OCR fallback failed ({type(exc).__name__}: {exc})"
        out, truncated = _truncate(text)
        if note:
            out = f"[{note}]\n{out}" if out.strip() else f"[{note}]"
        return ToolResult(
            ok=True,
            output=out,
            data={
                "path": str(path),
                "chars": len(text),
                "truncated": truncated,
                **({"note": note} if note else {}),
            },
        )


class WriteDocumentTool(Tool):
    name = "write_document"
    reversibility = Reversibility.REVERSIBLE  # TX-01: prior file bytes are captured
    description = (
        "Create a document inside the session workspace. The file type follows "
        "the path suffix (.docx/.xlsx/.pptx/.pdf/.csv/.html/.txt/.md), or the "
        "optional `kind` override. String content is markdown-aware: '# ' "
        "headings, -/1. lists, **bold**/*italic*, ``` code fences, | pipe | "
        "tables and '---' rules become REAL headings, lists, tables and slides "
        "in .docx/.pdf/.pptx/.html (plain text still works everywhere). A list "
        "of rows (list[list]) writes spreadsheet/CSV rows, and .xlsx also "
        "accepts {'sheets': {name: rows}} for a multi-sheet workbook."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "content": {
                "description": (
                    "A string (markdown renders as rich formatting; plain text "
                    "becomes paragraphs/lines), a list of rows for "
                    "spreadsheet/CSV output, or {'sheets': {name: rows}} for a "
                    "multi-sheet .xlsx."
                )
            },
            "kind": {
                "type": "string",
                "description": "Optional format override, e.g. 'pdf' or 'docx'.",
            },
        },
        "required": ["path", "content"],
    }

    async def capture_undo(
        self, args: dict[str, Any], ctx: ToolContext
    ) -> "dict[str, Any] | None":
        """Snapshot the inverse: prior RAW bytes when overwriting an existing
        document (``file_restore``), else a delete of the created file
        (``file_delete``). Documents are binary, so ``mode`` is ``raw`` and no
        ``post_sha256`` is predicted (the writer's exact bytes aren't known before
        it runs) — the pre-image still restores the prior file faithfully."""
        try:
            target = safe_path(ctx.workspace, args["path"])
        except Exception:
            return None
        if target.is_file():
            try:
                prior = target.read_bytes()
            except OSError:
                return None
            return make_file_descriptor(
                ctx.config.home,
                kind="file_restore",
                path=args["path"],
                mode="raw",
                prior_bytes=prior,
                pre_sha256=sha256_bytes(prior),
            )
        return make_file_descriptor(
            ctx.config.home,
            kind="file_delete",
            path=args["path"],
            mode="raw",
        )

    async def revert(self, undo: dict[str, Any], ctx: ToolContext) -> ToolResult:
        return await revert_workspace_file(undo, ctx)

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        try:
            target = safe_path(ctx.workspace, args["path"])
            out = write_document(target, args["content"], kind=args.get("kind"))
        except Exception as exc:
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        rel = str(out.relative_to(Path(ctx.workspace).resolve())).replace("\\", "/")
        size = out.stat().st_size
        return ToolResult(
            ok=True,
            output=f"wrote {size} bytes to {rel}",
            data={"path": rel, "bytes": size},
        )


class ExtractPdfTool(Tool):
    name = "extract_pdf"
    reversibility = Reversibility.READONLY  # extraction has no side effect
    returns_untrusted_content = True  # a PDF can carry planted instructions
    description = "Extract the text of a PDF file (absolute or workspace-relative path)."
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "page_range": {
                "type": "string",
                "description": (
                    "Optional 1-based page slice, e.g. '2', '1-3', '2-', "
                    "'1,4-6' — read only part of a large PDF."
                ),
            },
        },
        "required": ["path"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        path = _resolve_read_path(args["path"], ctx)
        if path.suffix.lower() != ".pdf":
            return ToolResult(ok=False, error=f"not a PDF file: {args['path']}")
        allowed, reason = fs_read_ok(path)
        if not allowed:
            return ToolResult(ok=False, error=reason)
        try:
            text = await asyncio.to_thread(  # CPU-bound parse off the loop
                extract_text, path, page_range=args.get("page_range")
            )
        except Exception as exc:
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        out, truncated = _truncate(text)
        return ToolResult(
            ok=True,
            output=out,
            data={"path": str(path), "chars": len(text), "truncated": truncated},
        )


#: Formats where a conversion must preserve real rows/cells, not flattened text.
_TABULAR = {".csv", ".xlsx"}


def _load_for_conversion(source: Path, src_suffix: str, tgt_suffix: str) -> Any:
    """Read ``source`` as the richest content shape the target can accept."""
    if src_suffix in _TABULAR and tgt_suffix in _TABULAR:
        if src_suffix == ".csv":  # real csv parsing, not text lines
            with open(source, newline="", encoding="utf-8", errors="replace") as f:
                return [list(row) for row in csv.reader(f)]
        return _xlsx_content(source, keep_sheets=(tgt_suffix == ".xlsx"))
    # PDF/office/HTML -> Markdown: preserve real structure (headings, lists,
    # tables) instead of flattening to plain text. Written verbatim into the
    # .md by write_document's text writer. Falls back to extract_text inside
    # document_to_markdown if markitdown can't handle the source.
    if tgt_suffix == ".md" and src_suffix in MARKITDOWN_SUFFIXES:
        return document_to_markdown(source)
    return extract_text(source)


def _xlsx_content(source: Path, *, keep_sheets: bool) -> Any:
    """Re-extract real rows from a workbook via openpyxl (not flattened text)."""
    from openpyxl import load_workbook

    # xlsx->xlsx keeps formulas (data_only=False); xlsx->csv wants cached values.
    wb = load_workbook(
        filename=str(source), read_only=True, data_only=not keep_sheets
    )
    try:
        sheets = {
            ws.title: [
                ["" if c is None else c for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            for ws in wb.worksheets
        }
    finally:
        wb.close()
    if keep_sheets and len(sheets) > 1:
        return {"sheets": sheets}
    return [row for rows in sheets.values() for row in rows]


class ConvertDocumentTool(Tool):
    name = "convert_document"
    description = (
        "Convert a document from one format to another: read any supported "
        "source (PDF, .docx, .xlsx, .pptx, .csv, text/code) and re-write it as "
        "any supported target (.docx/.xlsx/.pptx/.pdf/.csv/.html/.txt/.md). "
        "csv<->xlsx conversions preserve real rows and cells; other sources "
        "are extracted to text and rendered with markdown-aware formatting. "
        "The source may be ANY local path (absolute or workspace-relative); "
        "the target is created inside the session workspace."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "source": {
                "type": "string",
                "description": "Path of the document to read.",
            },
            "target": {
                "type": "string",
                "description": (
                    "Path of the document to create; its suffix picks the "
                    "output format."
                ),
            },
        },
        "required": ["source", "target"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        source = _resolve_read_path(args["source"], ctx)
        src_suffix = source.suffix.lower()
        if src_suffix not in SUPPORTED_READ:
            return ToolResult(
                ok=False,
                error=(
                    f"cannot read {src_suffix or source.name!r} — supported "
                    f"source formats: {', '.join(sorted(SUPPORTED_READ))}"
                ),
            )
        allowed, reason = fs_read_ok(source)
        if not allowed:
            return ToolResult(ok=False, error=reason)
        try:
            target = safe_path(ctx.workspace, args["target"])
        except Exception as exc:
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        tgt_suffix = target.suffix.lower()
        if tgt_suffix not in SUPPORTED_WRITE:
            return ToolResult(
                ok=False,
                error=(
                    f"cannot write {tgt_suffix or target.name!r} — supported "
                    f"target formats: {', '.join(sorted(SUPPORTED_WRITE))}"
                ),
            )
        try:
            content = await asyncio.to_thread(  # CPU-bound parse off the loop
                _load_for_conversion, source, src_suffix, tgt_suffix
            )
            out = await asyncio.to_thread(write_document, target, content)
        except Exception as exc:  # converting real files must never crash the runtime
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        rel = str(out.relative_to(Path(ctx.workspace).resolve())).replace("\\", "/")
        size = out.stat().st_size
        return ToolResult(
            ok=True,
            output=f"converted {source.name} -> {rel} ({size} bytes)",
            data={"source": str(source), "path": rel, "bytes": size},
        )


class RedactPiiTool(Tool):
    name = "redact_pii"
    reversibility = Reversibility.REVERSIBLE  # only the NEW output file to undo
    description = (
        "Redact PII from a document and write a NEW file in the SAME format — "
        "the original is never modified. Detects SSN/ITIN/EIN, emails, phones, "
        "credit cards, labeled account numbers, dates of birth, street "
        "addresses, and IPs; pass names or other exact strings you spotted "
        "while reading via extra_terms. Styles: 'black' (same-length █ blocks), "
        "'label' ([SSN]-style tags), 'remove' (deleted). Formats: .docx/.xlsx/"
        ".pptx keep their styling; text formats rewrite in place; .pdf is "
        "REBUILT from extracted text (content truly removed, layout "
        "approximate). Never quote the detected PII values back in your reply."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "Source document (absolute, or workspace-relative)",
            },
            "style": {
                "type": "string",
                "enum": ["black", "label", "remove"],
                "description": "black = █ blocks (default), label = [SSN] tags, remove = delete",
            },
            "extra_terms": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    "Exact strings to also redact (person names, employers, "
                    "spouse/dependent names…) — case-insensitive"
                ),
            },
            "categories": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    "Optional subset to redact: ssn, itin, ein, email, phone, "
                    "credit_card, bank_account, dob, address, ip, custom "
                    "(default: all)"
                ),
            },
            "output_path": {
                "type": "string",
                "description": (
                    "Workspace-relative output file (default: "
                    "<name>.redacted.<ext> beside the source when it lives in "
                    "the workspace, else in the workspace root)"
                ),
            },
        },
        "required": ["path"],
    }

    def _output_target(self, source: Path, args: dict[str, Any], ctx: ToolContext) -> Path:
        raw = str(args.get("output_path") or "").strip()
        if raw:
            return safe_path(ctx.workspace, raw)
        redacted_name = f"{source.stem}.redacted{source.suffix}"
        ws = Path(ctx.workspace).resolve()
        try:
            source.resolve().relative_to(ws)
            return safe_path(ctx.workspace, str(source.parent / redacted_name))
        except ValueError:
            # Source lives outside the workspace — the redacted copy lands in
            # the workspace root (writes are always workspace-confined).
            return safe_path(ctx.workspace, redacted_name)

    async def capture_undo(
        self, args: dict[str, Any], ctx: ToolContext
    ) -> "dict[str, Any] | None":
        """The ONLY side effect is the new output file — undo deletes it (or
        restores prior bytes if the target already existed)."""
        try:
            source = _resolve_read_path(str(args.get("path", "")), ctx)
            target = self._output_target(source, args, ctx)
            rel = str(target.relative_to(Path(ctx.workspace).resolve())).replace("\\", "/")
        except Exception:
            return None
        if target.is_file():
            try:
                prior = target.read_bytes()
            except OSError:
                return None
            return make_file_descriptor(
                ctx.config.home,
                kind="file_restore",
                path=rel,
                mode="raw",
                prior_bytes=prior,
                pre_sha256=sha256_bytes(prior),
            )
        return make_file_descriptor(ctx.config.home, kind="file_delete", path=rel, mode="raw")

    async def revert(self, undo: dict[str, Any], ctx: ToolContext) -> ToolResult:
        return await revert_workspace_file(undo, ctx)

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        from .redact import ALL_CATEGORIES, STYLES, redact_file

        source = _resolve_read_path(str(args.get("path", "")), ctx)
        allowed, reason = fs_read_ok(str(source))
        if not allowed:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if not source.is_file():
            return ToolResult(ok=False, error=f"not a file: {args.get('path')}")
        style = str(args.get("style") or "black").strip().lower()
        if style not in STYLES:
            return ToolResult(
                ok=False, error=f"unknown style {style!r} — use black, label, or remove"
            )
        cats_raw = args.get("categories") or []
        categories = None
        if cats_raw:
            categories = {str(c).strip().lower() for c in cats_raw if str(c).strip()}
            unknown = categories - ALL_CATEGORIES
            if unknown:
                return ToolResult(
                    ok=False,
                    error=(
                        f"unknown categories: {', '.join(sorted(unknown))} — "
                        f"valid: {', '.join(sorted(ALL_CATEGORIES))}"
                    ),
                )
            categories |= {"custom"}  # extra_terms always apply when provided
        extra_terms = [
            str(t) for t in (args.get("extra_terms") or []) if str(t).strip()
        ]
        try:
            target = self._output_target(source, args, ctx)
            if target.resolve() == source.resolve():
                return ToolResult(
                    ok=False,
                    error="output_path must differ from the source — the original is never overwritten",
                )
            target.parent.mkdir(parents=True, exist_ok=True)
            counts, note = await asyncio.to_thread(  # CPU-bound rewrite off the loop
                redact_file,
                source,
                target,
                style=style,
                extra_terms=extra_terms,
                categories=categories,
            )
        except Exception as exc:
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        rel = str(target.relative_to(Path(ctx.workspace).resolve())).replace("\\", "/")
        total = sum(counts.values())
        summary = (
            ", ".join(f"{cat}: {n}" for cat, n in sorted(counts.items()))
            if counts
            else "no PII found (output is an identical copy)"
        )
        return ToolResult(
            ok=True,
            output=(
                f"redacted {total} PII item(s) [{summary}] -> {rel} (style: {style})"
                + (f"\nNote: {note}" if note else "")
                + "\nThe original file was not modified."
            ),
            data={
                "path": rel,
                "source": str(source),
                "style": style,
                "counts": counts,
                "total": total,
                "note": note,
            },
        )


def document_tools(router_resolver: "Any | None" = None) -> list[Tool]:
    """Build the document tools. ``router_resolver`` (() -> ModelRouter) is
    optional and powers the scanned-PDF OCR fallback in ``read_document``;
    without it the tools behave exactly as before (no platform dependency)."""
    from .excel_tools import excel_tools

    return [
        ReadDocumentTool(router_resolver),
        WriteDocumentTool(),
        ExtractPdfTool(),
        ConvertDocumentTool(),
        ListFolderTool(),
        RedactPiiTool(),
        *excel_tools(),
    ]
