"""Deterministic document QA lint (v1.134.0).

Nothing in the pipeline ever LOOKED at a produced document — an empty section,
a zero-row table, or a sheet full of ``#REF!`` shipped silently. This module is
the deterministic half of the self-review loop: pure structural checks over a
rendered ``.docx``/``.xlsx``, no LLM anywhere, so a finding is a fact about the
file, never an opinion.

* :func:`lint_docx` — python-docx: empty document; headings with no body before
  the next same-or-higher heading; tables without data rows; a conservatively
  flagged truncated final paragraph; literal "1." → "3." numbering gaps.
* :func:`lint_xlsx` — openpyxl: empty / header-only sheets; headed columns with
  no data; Excel error literals (``#REF!`` family); formulas referencing past
  the sheet's used range; digit-strings in majority-numeric columns; charts
  whose ranges contain no data.
* :func:`lint_document` — suffix/kind dispatch; ``None`` for formats without a
  linter (a ``.txt`` has no structure to lint).

Every check returns findings shaped ``{"code", "where", "detail", "severity"}``
inside ``{"findings": [...], "ok": bool}`` where ``ok`` means "no error-severity
findings". Severity policy: ``error`` = the deliverable is materially broken
(nothing where content was promised, literal error values, charts over
nothing); ``warn`` = heuristic or sloppy-but-functional (truncated-tail guess,
numbering gaps, formulas past the used range, numbers stored as text) — the
refinement round in :mod:`.batch` spends an LLM call on errors only.

Contract: linting NEVER raises. An unreadable/corrupt file is itself a single
error finding (``docx-unreadable``/``xlsx-unreadable``) — a QA pass that
crashes the pipeline it guards would be worse than no QA at all.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

#: The truncated-tail heuristic only fires on documents at least this large.
#: A SHORT document ending without terminal punctuation is a style choice; a
#: large one stopping mid-sentence is the signature of a model reply clipped at
#: a token/size limit. Deliberately conservative — a false positive burns a
#: refinement round on a healthy document.
TRUNCATION_MIN_DOC_CHARS = 6_000

#: ...and the final paragraph must be real prose, not a short label/sign-off.
_TRUNC_MIN_TAIL_CHARS = 40

#: Per-code, per-sheet cap on per-cell findings: a sheet with 5,000 ``#REF!``
#: cells reports the first few plus an honest aggregate count, never a flood
#: (findings travel through tool output and LLM refinement prompts).
_PER_CODE_CAP = 8

#: Excel's literal error values — a model writing these as cell VALUES shipped
#: a visibly broken workbook. (The full standard set, not just the common four:
#: one code either way, and #N/A / #NUM! / #NULL! are no less broken.)
_XL_ERRORS = frozenset(
    {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}
)

_HEADING_STYLE_RX = re.compile(r"[Hh]eading (\d+)")
#: Literal numbered-list paragraphs ("1. text" / "1) text"). docx list STYLES
#: carry no literal digits (Word numbers them), so this only sees plain-text
#: numbering — exactly where a model can skip from "1." to "3.".
_NUMBERED_RX = re.compile(r"\s*(\d{1,3})[.)]\s+\S")

#: Pure digit-string (int/decimal). Used for the number-as-text check.
_DIGIT_STR_RX = re.compile(r"-?\d+(?:\.\d+)?")

#: Formula-reference scan, applied after quoted strings and sheet-qualified
#: refs are stripped: A1 cells/ranges with EXPLICIT row numbers only (so
#: whole-column ``A:A`` refs never match), not preceded by ``!`` (sheet-
#: qualified) and not followed by ``(`` (function names like LOG10).
_QUOTED_RX = re.compile(r'"[^"]*"')
_SHEET_QUALIFIED_RX = re.compile(
    r"(?:'[^']*'|[A-Za-z0-9_.]+)!\$?[A-Za-z]{1,3}\$?\d{1,7}"
    r"(?::\$?[A-Za-z]{1,3}\$?\d{1,7})?"
)
_REF_RX = re.compile(
    r"(?<![A-Za-z0-9_$:!])\$?([A-Za-z]{1,3})\$?(\d{1,7})"
    r"(?::\$?([A-Za-z]{1,3})\$?(\d{1,7}))?(?![A-Za-z0-9_(])"
)


def _finding(code: str, where: str, detail: str, severity: str) -> dict[str, str]:
    return {"code": code, "where": where, "detail": detail, "severity": severity}


def _result(findings: list[dict[str, str]]) -> dict[str, Any]:
    return {
        "findings": findings,
        "ok": not any(f["severity"] == "error" for f in findings),
    }


def _extend_capped(
    findings: list[dict[str, str]], batch: list[dict[str, str]], where: str
) -> None:
    """Append ``batch`` (all one code) honoring :data:`_PER_CODE_CAP` — the
    overflow collapses into one aggregate finding, never silence."""
    findings.extend(batch[:_PER_CODE_CAP])
    if len(batch) > _PER_CODE_CAP:
        first = batch[0]
        findings.append(
            _finding(
                first["code"],
                where,
                f"...and {len(batch) - _PER_CODE_CAP} more like this",
                first["severity"],
            )
        )


# --- docx ----------------------------------------------------------------------


def lint_docx(path: str | Path) -> dict[str, Any]:
    """Deterministic structural lint of a ``.docx``. Never raises — an
    unreadable file is a single ``docx-unreadable`` error finding."""
    try:
        return _result(_docx_findings(Path(path)))
    except Exception as exc:  # noqa: BLE001 — QA must never crash what it guards
        return _result(
            [
                _finding(
                    "docx-unreadable",
                    Path(path).name,
                    f"cannot open as .docx — {type(exc).__name__}: {exc}",
                    "error",
                )
            ]
        )


def _heading_level(para: Any) -> "int | None":
    """1-4 for "Heading N" styles, None for body text (incl. style trouble)."""
    try:
        name = para.style.name or ""
    except Exception:  # noqa: BLE001 — a style-less paragraph is body text
        return None
    m = _HEADING_STYLE_RX.fullmatch(name)
    return int(m.group(1)) if m else None


def _style_name(para: Any) -> str:
    try:
        return para.style.name or ""
    except Exception:  # noqa: BLE001
        return ""


def _docx_findings(path: Path) -> list[dict[str, str]]:
    import docx
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = docx.Document(str(path))
    # Walk the body's DIRECT children in document order — doc.paragraphs loses
    # the paragraph/table interleaving the empty-section check depends on.
    items: list[Any] = []
    for child in doc.element.body.iterchildren():
        if child.tag == qn("w:p"):
            items.append(Paragraph(child, doc))
        elif child.tag == qn("w:tbl"):
            items.append(Table(child, doc))

    findings: list[dict[str, str]] = []

    # 1. Empty document: no non-whitespace paragraph and no table at all.
    has_text = any(
        isinstance(it, Paragraph) and it.text.strip() for it in items
    )
    has_table = any(isinstance(it, Table) for it in items)
    if not has_text and not has_table:
        return [
            _finding(
                "docx-empty",
                path.name,
                "document has no non-whitespace content",
                "error",
            )
        ]

    # 2. Empty sections: a heading with no body (non-heading paragraph text or
    #    a table) before the next SAME-OR-HIGHER heading. Deeper subheadings
    #    neither end the section nor count as its body.
    heads = [
        (i, lvl, it.text.strip())
        for i, it in enumerate(items)
        if isinstance(it, Paragraph)
        and (lvl := _heading_level(it)) is not None
    ]
    for pos, (i, level, text) in enumerate(heads):
        end = len(items)
        for j, jl, _t in heads[pos + 1 :]:
            if jl <= level:
                end = j
                break
        body = False
        for it in items[i + 1 : end]:
            if isinstance(it, Table) or (
                isinstance(it, Paragraph)
                and _heading_level(it) is None
                and it.text.strip()
            ):
                body = True
                break
        if not body:
            findings.append(
                _finding(
                    "docx-empty-section",
                    f"heading {text or '(untitled)'!r}",
                    "no body content before the next same-or-higher heading",
                    "error",
                )
            )

    # 3. Tables with no data rows (a lone header row promises data it lacks).
    tbl_no = 0
    for it in items:
        if isinstance(it, Table):
            tbl_no += 1
            if len(it.rows) <= 1:
                findings.append(
                    _finding(
                        "docx-table-no-data",
                        f"table {tbl_no}",
                        f"{len(it.rows)} row(s) — a header row with no data rows",
                        "error",
                    )
                )

    # 4. Literal numbering gaps: "1." followed by "3." in consecutive numbered
    #    paragraphs. Any non-numbered paragraph (incl. blank) resets the run —
    #    conservative on purpose: restarts and interleaved prose never flag.
    prev: "int | None" = None
    for idx, it in enumerate(items):
        m = _NUMBERED_RX.match(it.text) if isinstance(it, Paragraph) else None
        if not m:
            prev = None
            continue
        n = int(m.group(1))
        if prev is not None and n > prev + 1:
            findings.append(
                _finding(
                    "docx-numbering-gap",
                    f"paragraph {idx + 1}",
                    f'item "{prev}." is followed by "{n}." — the numbering skips',
                    "warn",
                )
            )
        prev = n

    # 5. Truncated tail — CONSERVATIVE heuristic, all conditions required:
    #    the document's last content is a PROSE paragraph (not a heading, list
    #    item, or table), of real length, whose final character is a letter /
    #    comma / dash / opening bracket (i.e. mid-sentence, not merely missing
    #    a period), AND the document is large (TRUNCATION_MIN_DOC_CHARS) — the
    #    size boundary that makes "a model reply got clipped" the likely story.
    total_chars = sum(
        len(it.text) for it in items if isinstance(it, Paragraph)
    )
    tail = None
    for it in reversed(items):
        if isinstance(it, Table):
            break  # a document ending in a table is not a clipped reply
        if isinstance(it, Paragraph) and it.text.strip():
            tail = it
            break
    if tail is not None and total_chars >= TRUNCATION_MIN_DOC_CHARS:
        text = tail.text.strip()
        prose = _heading_level(tail) is None and not _style_name(tail).startswith(
            "List"
        )
        if (
            prose
            and len(text) >= _TRUNC_MIN_TAIL_CHARS
            and (text[-1].isalpha() or text[-1] in ",-–—([")
        ):
            findings.append(
                _finding(
                    "docx-truncated-tail",
                    f"final paragraph (...{text[-30:]!r})",
                    "large document ends mid-sentence — the source reply may "
                    "have been clipped at a size limit",
                    "warn",
                )
            )

    return findings


# --- xlsx ----------------------------------------------------------------------


def lint_xlsx(path: str | Path) -> dict[str, Any]:
    """Deterministic structural lint of an ``.xlsx``. Never raises — an
    unreadable file is a single ``xlsx-unreadable`` error finding."""
    try:
        return _result(_xlsx_findings(Path(path)))
    except Exception as exc:  # noqa: BLE001 — QA must never crash what it guards
        return _result(
            [
                _finding(
                    "xlsx-unreadable",
                    Path(path).name,
                    f"cannot open as .xlsx — {type(exc).__name__}: {exc}",
                    "error",
                )
            ]
        )


def _blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _xlsx_findings(path: Path) -> list[dict[str, str]]:
    import io

    from openpyxl import load_workbook

    # Default (non-data_only, non-read_only) load: formulas stay as their "="
    # strings for the reference scan, and charts are materialized on ws._charts.
    # Loaded from bytes because openpyxl rejects non-.xlsx EXTENSIONS on string
    # paths — but lint dispatch honors write_document's `kind` override, where
    # the suffix can lie (an .xlsx written to "report.bin" is still an .xlsx).
    wb = load_workbook(io.BytesIO(path.read_bytes()))
    try:
        findings: list[dict[str, str]] = []
        for ws in wb.worksheets:
            _sheet_findings(ws, wb, findings)
        return findings
    finally:
        wb.close()


def _sheet_findings(ws: Any, wb: Any, findings: list[dict[str, str]]) -> None:
    from openpyxl.utils import get_column_letter

    title = ws.title
    vals = [[c.value for c in row] for row in ws.iter_rows()]

    # 1. Completely empty sheet — charts can still (brokenly) sit on it.
    if not any(not _blank(v) for row in vals for v in row):
        findings.append(
            _finding(
                "xlsx-empty-sheet",
                f"sheet {title!r}",
                "sheet is completely empty",
                "error",
            )
        )
        _chart_findings(ws, wb, findings)
        return

    # Header detection mirrors the writer's own heuristic (_xlsx_fill): a
    # first row that is entirely non-formula strings reads as a header row —
    # but only with AT LEAST TWO filled cells. A lone string cell is a
    # note/title, not a promise of columns: the writer itself puts one-line
    # string content in a single cell, and flagging that as "header row only"
    # was a false positive on the app's own output.
    header = sum(1 for v in vals[0] if not _blank(v)) >= 2 and all(
        isinstance(v, str) and not v.startswith("=") for v in vals[0]
    )
    data_rows = vals[1:] if header else vals

    # 2. Header-only sheet: a header row promising data with 0 data rows.
    if header and not any(not _blank(v) for row in data_rows for v in row):
        findings.append(
            _finding(
                "xlsx-empty-sheet",
                f"sheet {title!r}",
                "header row only — 0 data rows",
                "error",
            )
        )
    elif header:
        # 3. Headed columns whose every data cell is empty (truncated output).
        for ci, head in enumerate(vals[0]):
            if _blank(head):
                continue
            if all(_blank(row[ci]) for row in data_rows):
                letter = get_column_letter(ci + 1)
                findings.append(
                    _finding(
                        "xlsx-empty-column",
                        f"'{title}'!{letter}",
                        f"column {letter} has header {str(head)!r} but every "
                        "data cell is empty",
                        "error",
                    )
                )

    # 4. Excel error literals stored as cell values.
    err_batch: list[dict[str, str]] = []
    for ri, row in enumerate(vals, start=1):
        for ci, v in enumerate(row, start=1):
            if isinstance(v, str) and v.strip() in _XL_ERRORS:
                err_batch.append(
                    _finding(
                        "xlsx-error-cell",
                        f"'{title}'!{get_column_letter(ci)}{ri}",
                        f"cell contains the Excel error literal {v.strip()}",
                        "error",
                    )
                )
    _extend_capped(findings, err_batch, f"sheet {title!r}")

    # 5. Formulas referencing beyond the sheet's used range. Same-sheet refs
    #    with explicit rows only — sheet-qualified and whole-column refs are
    #    skipped rather than guessed at (conservative by design).
    frm_batch: list[dict[str, str]] = []
    for ri, row in enumerate(vals, start=1):
        for ci, v in enumerate(row, start=1):
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            ref = _formula_ref_beyond(v, ws.max_row, ws.max_column)
            if ref:
                frm_batch.append(
                    _finding(
                        "xlsx-formula-out-of-range",
                        f"'{title}'!{get_column_letter(ci)}{ri}",
                        f"formula references {ref}, beyond the sheet's used "
                        f"range (A1:{get_column_letter(ws.max_column)}{ws.max_row})",
                        "warn",
                    )
                )
    _extend_capped(findings, frm_batch, f"sheet {title!r}")

    # 6. Numbers stored as text in majority-numeric columns. Conservative:
    #    >50% of the column's filled data cells must already be numeric, the
    #    cell must be a pure digit-string, and leading-zero identifiers
    #    ("007" — the writer's own keep-as-text convention) never flag.
    txt_batch: list[dict[str, str]] = []
    start_row = 2 if header else 1
    for ci in range(len(vals[0])):
        col = [row[ci] for row in data_rows]
        filled = [v for v in col if not _blank(v)]
        numeric = [
            v
            for v in filled
            if isinstance(v, (int, float)) and not isinstance(v, bool)
        ]
        if len(filled) < 2 or len(numeric) * 2 <= len(filled):
            continue
        for off, v in enumerate(col):
            if (
                isinstance(v, str)
                and _DIGIT_STR_RX.fullmatch(v.strip())
                and not _leading_zero_id(v.strip())
            ):
                txt_batch.append(
                    _finding(
                        "xlsx-number-as-text",
                        f"'{title}'!{get_column_letter(ci + 1)}{start_row + off}",
                        f"{v!r} is a number stored as text in a "
                        "majority-numeric column",
                        "warn",
                    )
                )
    _extend_capped(findings, txt_batch, f"sheet {title!r}")

    # 7. Charts whose ranges contain no data.
    _chart_findings(ws, wb, findings)


def _leading_zero_id(s: str) -> bool:
    t = s.lstrip("-")
    return len(t) > 1 and t[0] == "0" and not t.startswith("0.")


def _formula_ref_beyond(formula: str, max_row: int, max_col: int) -> "str | None":
    """First same-sheet A1 reference past (max_row, max_col), else None."""
    from openpyxl.utils import column_index_from_string

    body = _SHEET_QUALIFIED_RX.sub(" ", _QUOTED_RX.sub(" ", formula[1:]))
    for m in _REF_RX.finditer(body):
        cols = [column_index_from_string(m.group(1).upper())]
        rows = [int(m.group(2))]
        if m.group(3):
            cols.append(column_index_from_string(m.group(3).upper()))
            rows.append(int(m.group(4)))
        if max(rows) > max_row or max(cols) > max_col:
            return m.group(0)
    return None


def _chart_findings(ws: Any, wb: Any, findings: list[dict[str, str]]) -> None:
    """Flag charts whose series/category ranges contain no data at all.

    Partially filled ranges pass (a chart over B2:B100 with 10 values renders
    fine); unresolvable refs are skipped, not guessed at. ``ws._charts`` is
    openpyxl's own (private but stable) chart anchor list — there is no public
    read API for charts.
    """
    for ch_no, chart in enumerate(getattr(ws, "_charts", None) or [], start=1):
        for f in _chart_ref_strings(chart):
            target, bounds = _resolve_ref(f, wb, ws)
            if target is None or bounds is None:
                continue
            if _range_all_blank(target, bounds):
                findings.append(
                    _finding(
                        "xlsx-chart-empty-range",
                        f"chart {ch_no} on sheet {ws.title!r}",
                        f"chart range {f} contains no data",
                        "error",
                    )
                )


def _chart_ref_strings(chart: Any) -> list[str]:
    out: list[str] = []
    for ser in list(getattr(chart, "series", None) or []):
        for axis in (getattr(ser, "val", None), getattr(ser, "cat", None)):
            for kind in ("numRef", "strRef"):
                f = getattr(getattr(axis, kind, None), "f", None)
                if f:
                    out.append(str(f))
    return out


def _resolve_ref(
    f: str, wb: Any, default_ws: Any
) -> "tuple[Any, tuple[int, int, int, int] | None]":
    """('Sheet'!$B$2:$B$9 → (worksheet, (min_col, min_row, max_col, max_row));
    (None, None) when the sheet or range cannot be resolved."""
    from openpyxl.utils import range_boundaries

    s = str(f).lstrip("=").strip()
    if "!" in s:
        sheet, rng = s.rsplit("!", 1)
        sheet = sheet.strip().strip("'").replace("''", "'")
        ws = wb[sheet] if sheet in wb.sheetnames else None
    else:
        ws, rng = default_ws, s
    if ws is None:
        return None, None
    try:
        bounds = range_boundaries(rng.replace("$", "").upper())
    except Exception:  # noqa: BLE001 — malformed refs degrade, never crash
        return None, None
    if None in bounds:
        return None, None
    return ws, bounds  # (min_col, min_row, max_col, max_row)


def _range_all_blank(ws: Any, bounds: tuple[int, int, int, int]) -> bool:
    min_col, min_row, max_col, max_row = bounds
    if min_row > ws.max_row or min_col > ws.max_column:
        return True  # entirely outside the used range = entirely empty
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=min(max_row, ws.max_row),
        min_col=min_col,
        max_col=min(max_col, ws.max_column),
    ):
        for c in row:
            if not _blank(c.value):
                return False
    return True


# --- dispatch ------------------------------------------------------------------

_LINTERS = {".docx": lint_docx, ".xlsx": lint_xlsx}


def lint_document(path: str | Path, kind: "str | None" = None) -> "dict[str, Any] | None":
    """Lint ``path`` by suffix (or the ``kind`` override, mirroring
    ``write_document``'s dispatch). Returns None for formats without a linter —
    "not lintable" and "linted clean" must never look alike."""
    suffix = ("." + kind.lstrip(".")).lower() if kind else Path(path).suffix.lower()
    fn = _LINTERS.get(suffix)
    return fn(path) if fn else None
