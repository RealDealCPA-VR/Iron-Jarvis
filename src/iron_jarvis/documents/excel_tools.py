"""Excel parity tools: structured reads + IN-PLACE edits of real workbooks.

What made "Claude works with Excel" magic is not the model — it's the pair of
capabilities these tools provide: reading a workbook as STRUCTURE (sheets,
cells, formulas, ranges — not a text dump) and editing an EXISTING workbook
in place (set cells, write formulas, add sheets) while everything untouched
keeps its formatting. Any tool-capable model — local or frontier — drives
them identically, so a verified local endpoint gets full parity.

Honesty & safety:
* reads reach any policy-allowed path; edits are WORKSPACE-confined (the
  project folder in practice) via ``safe_path`` like every write tool;
* edits are TX-01 reversible (prior bytes captured — undo restores the exact
  workbook);
* openpyxl preserves cell formatting/formulas it doesn't touch, but exotic
  content (macros beyond .xlsm passthrough, some chart types) can be
  simplified — the tool says so rather than pretending otherwise.
"""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any

from ..core.fs_policy import fs_read_ok
from ..tools.base import Reversibility, Tool, ToolContext, ToolResult, safe_path
from ..tools.undo import make_file_descriptor, revert_workspace_file, sha256_bytes

_MAX_CELLS = 4000  # structured-read cap: keeps a huge sheet from flooding context


def _resolve_read_path(raw: str, ctx: ToolContext) -> Path:
    p = Path(raw)
    return p if p.is_absolute() else Path(ctx.workspace) / raw


def _read_workbook(path: Path, sheet: "str | None", cell_range: "str | None",
                   include_formulas: bool) -> dict[str, Any]:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries

    wb = load_workbook(str(path), data_only=not include_formulas)
    out: dict[str, Any] = {"sheets": wb.sheetnames}
    if sheet is None and cell_range is None:
        # Overview mode: names + dimensions per sheet.
        out["overview"] = [
            {"sheet": ws.title, "rows": ws.max_row, "cols": ws.max_column}
            for ws in wb.worksheets
        ]
        return out
    ws = wb[sheet] if sheet else wb.active
    out["sheet"] = ws.title
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
    cells = 0
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        r: list[Any] = []
        for c in row:
            v = c.value
            r.append(str(v) if v is not None and not isinstance(v, (int, float, bool)) else v)
            cells += 1
        rows.append(r)
        if cells >= _MAX_CELLS:
            out["truncated"] = True
            break
    out["rows"] = rows
    return out


class ExcelReadTool(Tool):
    name = "excel_read"
    reversibility = Reversibility.READONLY
    returns_untrusted_content = True  # spreadsheet text can carry planted instructions
    description = (
        "Read an Excel workbook as STRUCTURE, not text: with only a path you "
        "get the sheet list + dimensions; add `sheet` and/or `range` (e.g. "
        "'A1:D20') for cell values as rows. Set include_formulas=true to see "
        "formulas instead of computed values. Any policy-allowed path."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string", "description": "Sheet name (default: active)"},
            "range": {"type": "string", "description": "A1-style range, e.g. B2:F30"},
            "include_formulas": {"type": "boolean"},
        },
        "required": ["path"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        path = _resolve_read_path(str(args.get("path", "")), ctx)
        ok, reason = fs_read_ok(str(path))
        if not ok:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"not an Excel workbook: {path.name}")
        try:
            data = await asyncio.to_thread(
                _read_workbook, path, args.get("sheet"), args.get("range"),
                bool(args.get("include_formulas")),
            )
        except Exception as exc:  # noqa: BLE001 — real files must not crash the runtime
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        import json

        return ToolResult(ok=True, output=json.dumps(data, ensure_ascii=False), data=data)


def _apply_edits(path: Path, sheet: "str | None", edits: list[dict[str, Any]],
                 add_sheets: list[str]) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path))  # keep_vba not needed for .xlsx; .xlsm noted
    for name in add_sheets:
        if name not in wb.sheetnames:
            wb.create_sheet(title=name)
    applied = 0
    for e in edits:
        target = wb[str(e.get("sheet"))] if e.get("sheet") else (
            wb[sheet] if sheet else wb.active
        )
        cell = str(e.get("cell", "")).strip()
        if not cell:
            raise ValueError("every edit needs a cell (e.g. 'B2')")
        if "formula" in e:
            target[cell] = str(e["formula"])
        else:
            target[cell] = e.get("value")
        applied += 1
    wb.save(str(path))
    return {"applied": applied, "sheets": wb.sheetnames}


class ExcelEditTool(Tool):
    name = "excel_edit"
    reversibility = Reversibility.REVERSIBLE  # TX-01: prior workbook bytes captured
    description = (
        "Edit an EXISTING Excel workbook IN PLACE inside the workspace: set "
        "cell values and formulas ({cell:'B2', value:…} or {cell:'C3', "
        "formula:'=SUM(A1:A10)'}, optional per-edit sheet), and add sheets. "
        "Untouched cells keep their formatting/formulas. Undoable. Note: "
        "exotic content (some charts, macros) may be simplified by the "
        "editor — mention it when the workbook looks complex."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Workspace-relative workbook path"},
            "sheet": {"type": "string", "description": "Default sheet for edits"},
            "edits": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "cell": {"type": "string"},
                        "value": {},
                        "formula": {"type": "string"},
                        "sheet": {"type": "string"},
                    },
                    "required": ["cell"],
                },
            },
            "add_sheets": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["path"],
    }

    async def capture_undo(self, args: dict[str, Any], ctx: ToolContext) -> "dict[str, Any] | None":
        try:
            target = safe_path(ctx.workspace, args["path"])
        except Exception:
            return None
        if not target.is_file():
            return None
        try:
            prior = target.read_bytes()
        except OSError:
            return None
        return make_file_descriptor(
            ctx.config.home, kind="file_restore", path=args["path"],
            mode="raw", prior_bytes=prior, pre_sha256=sha256_bytes(prior),
        )

    async def revert(self, undo: dict[str, Any], ctx: ToolContext) -> ToolResult:
        return await revert_workspace_file(undo, ctx)

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        try:
            target = safe_path(ctx.workspace, args["path"])
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=str(exc))
        if not target.is_file():
            return ToolResult(
                ok=False,
                error=f"no such workbook in the workspace: {args.get('path')} "
                      "(use write_document to create one first)",
            )
        if target.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"not an Excel workbook: {target.name}")
        edits = [e for e in (args.get("edits") or []) if isinstance(e, dict)]
        add_sheets = [str(s) for s in (args.get("add_sheets") or []) if str(s).strip()]
        if not edits and not add_sheets:
            return ToolResult(ok=False, error="nothing to do — pass edits and/or add_sheets")
        try:
            result = await asyncio.to_thread(
                _apply_edits, target, args.get("sheet"), edits, add_sheets
            )
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        rel = str(target.relative_to(Path(ctx.workspace).resolve())).replace("\\", "/")
        return ToolResult(
            ok=True,
            output=f"applied {result['applied']} edit(s) to {rel} "
                   f"(sheets: {', '.join(result['sheets'])})",
            data={"path": rel, **result},
        )


# --------------------------------------------------------------------------- #
# Engine-computed analysis (v1.89.0): profile + query. Local models hallucinate
# most on ARITHMETIC — so numbers must come from the engine, never the model.
# Both tools return compact text (headers + exact figures), sized for small
# context windows. Pure openpyxl row iteration — no pandas dependency.
# --------------------------------------------------------------------------- #

_SCAN_ROW_CAP = 50_000  # rows scanned per query — bounded, with an honest flag


def _cell_str(v: Any) -> str:
    return "" if v is None else str(v)


def _load_table(path: Path, sheet: "str | None") -> tuple[str, list[str], list[list[Any]], bool]:
    """(sheet_title, headers, data_rows, truncated) — headers = first row."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers: list[str] = []
    rows: list[list[Any]] = []
    truncated = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [_cell_str(v).strip() for v in row]
            continue
        if len(rows) >= _SCAN_ROW_CAP:
            truncated = True
            break
        rows.append(list(row))
    title = ws.title
    wb.close()
    return title, headers, rows, truncated


def _col_index(column: str, headers: list[str]) -> int:
    """Resolve a column by HEADER NAME (case-insensitive) or Excel letter."""
    want = (column or "").strip()
    if not want:
        raise ValueError("a column is required (header name or letter like 'B')")
    lowered = [h.lower() for h in headers]
    if want.lower() in lowered:
        return lowered.index(want.lower())
    if want.isalpha() and len(want) <= 3:  # Excel letter(s)
        from openpyxl.utils import column_index_from_string

        return column_index_from_string(want.upper()) - 1
    raise ValueError(
        f"no column {want!r} — headers are: {', '.join(h for h in headers if h) or '(none)'}"
    )


def _as_number(v: Any) -> "float | None":
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


_WHERE_OPS = ("eq", "ne", "contains", "gt", "lt", "ge", "le")


def _row_matches(row: list[Any], cond: dict[str, Any], headers: list[str]) -> bool:
    idx = _col_index(str(cond.get("column", "")), headers)
    cell = row[idx] if idx < len(row) else None
    op = str(cond.get("op", "eq")).lower()
    want = cond.get("value")
    if op in ("gt", "lt", "ge", "le"):
        a, b = _as_number(cell), _as_number(want)
        if a is None or b is None:
            return False
        return {"gt": a > b, "lt": a < b, "ge": a >= b, "le": a <= b}[op]
    a_s, b_s = _cell_str(cell).strip().lower(), _cell_str(want).strip().lower()
    if op == "eq":
        return a_s == b_s
    if op == "ne":
        return a_s != b_s
    if op == "contains":
        return b_s in a_s
    raise ValueError(f"unknown where op {op!r} — use one of {_WHERE_OPS}")


class ExcelProfileTool(Tool):
    name = "excel_profile"
    reversibility = Reversibility.READONLY
    returns_untrusted_content = True  # sheet text can carry planted instructions
    description = (
        "Orient in an Excel workbook CHEAPLY: every sheet's name, row/column "
        "counts, headers, and one sample row — a compact map, not a data dump. "
        "Call this FIRST, then excel_query for figures. Any policy-allowed "
        "path (local, network share, or tailnet folder)."
    )
    input_schema = {
        "type": "object",
        "properties": {"path": {"type": "string"}},
        "required": ["path"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        path = _resolve_read_path(str(args.get("path", "")), ctx)
        ok, reason = fs_read_ok(str(path))
        if not ok:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"not an Excel workbook: {path.name}")

        def _profile() -> dict[str, Any]:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), data_only=True, read_only=True)
            sheets = []
            for ws in wb.worksheets:
                first = next(ws.iter_rows(values_only=True, max_row=1), tuple())
                second = next(
                    ws.iter_rows(values_only=True, min_row=2, max_row=2), tuple()
                )
                sheets.append({
                    "sheet": ws.title,
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                    "headers": [_cell_str(v).strip() for v in first[:24]],
                    "sample": [_cell_str(v)[:40] for v in second[:24]],
                })
            wb.close()
            return {"sheets": sheets}

        try:
            data = await asyncio.to_thread(_profile)
        except Exception as exc:  # noqa: BLE001 — real files must not crash the runtime
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        lines = [f"{path.name} — {len(data['sheets'])} sheet(s)"]
        for s in data["sheets"]:
            heads = ", ".join(h for h in s["headers"] if h) or "(no header row)"
            lines.append(f"- {s['sheet']}: {s['rows']}x{s['cols']} | headers: {heads}")
        return ToolResult(ok=True, output="\n".join(lines), data=data)


class ExcelQueryTool(Tool):
    name = "excel_query"
    reversibility = Reversibility.READONLY
    returns_untrusted_content = True  # spreadsheet text can carry planted instructions
    description = (
        "Compute over an Excel sheet with the ENGINE — exact numbers, never "
        "mental math: op sum/avg/min/max/count on a column, group (group_by + "
        "agg per group), or filter (matching rows). Columns by header name or "
        "letter; optional `where` conditions (eq/ne/contains/gt/lt/ge/le) "
        "combine with AND. ALWAYS use this for figures — report its results "
        "exactly. Any policy-allowed path."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string", "description": "Sheet name (default: active)"},
            "op": {
                "type": "string",
                "enum": ["sum", "avg", "min", "max", "count", "group", "filter"],
            },
            "column": {"type": "string", "description": "Target column (aggregates)"},
            "group_by": {"type": "string", "description": "Grouping column (op=group)"},
            "agg": {
                "type": "string",
                "enum": ["sum", "avg", "count"],
                "description": "Per-group aggregate (op=group, default sum)",
            },
            "where": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "column": {"type": "string"},
                        "op": {"type": "string", "enum": list(_WHERE_OPS)},
                        "value": {},
                    },
                    "required": ["column"],
                },
            },
            "limit": {"type": "integer", "description": "Rows/groups returned (default 20)"},
        },
        "required": ["path", "op"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        path = _resolve_read_path(str(args.get("path", "")), ctx)
        ok, reason = fs_read_ok(str(path))
        if not ok:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"not an Excel workbook: {path.name}")
        op = str(args.get("op", "")).lower()
        limit = max(1, min(int(args.get("limit") or 20), 50))
        try:
            title, headers, rows, truncated = await asyncio.to_thread(
                _load_table, path, args.get("sheet")
            )
            conds = [c for c in (args.get("where") or []) if isinstance(c, dict)]
            if conds:
                rows = [r for r in rows if all(_row_matches(r, c, headers) for c in conds)]

            if op == "filter":
                shown = rows[:limit]
                head = " | ".join(h or "·" for h in headers)
                body = "\n".join(
                    " | ".join(_cell_str(v)[:40] for v in r) for r in shown
                )
                note = f" (showing {len(shown)} of {len(rows)} matching rows)"
                out = f"{title}: {len(rows)} matching row(s){note}\n{head}\n{body}"
                return ToolResult(ok=True, output=out, data={
                    "sheet": title, "matches": len(rows), "headers": headers,
                    "rows": [[_cell_str(v) for v in r] for r in shown],
                    "scan_truncated": truncated,
                })

            if op == "group":
                gi = _col_index(str(args.get("group_by", "")), headers)
                agg = str(args.get("agg") or "sum").lower()
                ci = _col_index(str(args.get("column", "")), headers) if agg != "count" else -1
                groups: dict[str, list[float]] = {}
                counts: dict[str, int] = {}
                for r in rows:
                    key = _cell_str(r[gi] if gi < len(r) else None).strip() or "(blank)"
                    counts[key] = counts.get(key, 0) + 1
                    if ci >= 0:
                        n = _as_number(r[ci] if ci < len(r) else None)
                        if n is not None:
                            groups.setdefault(key, []).append(n)
                if agg == "count":
                    ranked = sorted(counts.items(), key=lambda kv: -kv[1])[:limit]
                    result = [{"group": k, "count": v} for k, v in ranked]
                    body = "\n".join(f"- {k}: {v}" for k, v in ranked)
                else:
                    scored = {
                        k: (sum(v) if agg == "sum" else sum(v) / len(v))
                        for k, v in groups.items() if v
                    }
                    ranked2 = sorted(scored.items(), key=lambda kv: -kv[1])[:limit]
                    result = [{"group": k, agg: v, "count": counts.get(k, 0)}
                              for k, v in ranked2]
                    body = "\n".join(f"- {k}: {v:,.2f}" for k, v in ranked2)
                out = f"{title}: {agg} by {args.get('group_by')} ({len(rows)} rows)\n{body}"
                return ToolResult(ok=True, output=out, data={
                    "sheet": title, "groups": result, "rows_scanned": len(rows),
                    "scan_truncated": truncated,
                })

            # Column aggregates: sum / avg / min / max / count.
            ci = _col_index(str(args.get("column", "")), headers)
            nums = [n for r in rows
                    if (n := _as_number(r[ci] if ci < len(r) else None)) is not None]
            nonempty = sum(1 for r in rows if _cell_str(r[ci] if ci < len(r) else None).strip())
            if op == "count":
                value: float = float(nonempty)
            elif not nums:
                return ToolResult(
                    ok=False,
                    error=f"column {args.get('column')!r} has no numeric values"
                          f" in the {len(rows)} row(s) considered",
                )
            elif op == "sum":
                value = sum(nums)
            elif op == "avg":
                value = sum(nums) / len(nums)
            elif op == "min":
                value = min(nums)
            elif op == "max":
                value = max(nums)
            else:
                return ToolResult(ok=False, error=f"unknown op {op!r}")
            skipped = nonempty - len(nums) if op != "count" else 0
            extra = f", {skipped} non-numeric skipped" if skipped > 0 else ""
            trunc = " [scan capped — figures cover the first 50k rows]" if truncated else ""
            out = (
                f"{title}: {op.upper()} of {args.get('column')} = {value:,.2f} "
                f"({len(rows)} row(s) considered{extra}){trunc}"
            )
            return ToolResult(ok=True, output=out, data={
                "sheet": title, "op": op, "column": args.get("column"),
                "value": value, "rows_considered": len(rows),
                "numeric_values": len(nums), "scan_truncated": truncated,
            })
        except ValueError as exc:
            return ToolResult(ok=False, error=str(exc))
        except Exception as exc:  # noqa: BLE001 — real files must not crash the runtime
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")


# --------------------------------------------------------------------------- #
# Formula intelligence (v1.90.0): derive → CHECK → reproduce → keep current.
# The evaluator (documents/excel_formula.py) computes a formula against the
# real sheet so a written formula is VALIDATED, not assumed; sheet specs
# capture formulas + formatting for faithful reproduction; the accounts diff
# keeps label-anchored formulas honest when rows come and go.
# --------------------------------------------------------------------------- #


class ExcelFormulaCheckTool(Tool):
    name = "excel_formula_check"
    reversibility = Reversibility.READONLY
    returns_untrusted_content = True  # sheet text can carry planted instructions
    description = (
        "VALIDATE a formula against the real workbook before trusting or "
        "writing it: computes the formula (SUM/AVERAGE/MIN/MAX/COUNT/ABS/"
        "ROUND/IF/SUBTOTAL, cell refs + ranges + arithmetic) over the sheet's "
        "actual values and, when `expected` or `expect_cell` is given, "
        "compares within `tolerance`. Pass `cell` instead of `formula` to "
        "check the formula already stored in that cell. Workflow: derive a "
        "formula → check it here → only then excel_edit it in. Any "
        "policy-allowed path."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string", "description": "Sheet name (default: active)"},
            "formula": {"type": "string", "description": "Formula to evaluate, e.g. '=SUM(B2:B9)'"},
            "cell": {"type": "string", "description": "Check the formula stored IN this cell"},
            "expected": {"type": "number", "description": "Expected numeric result"},
            "expect_cell": {"type": "string", "description": "Cell whose stored value the result should equal"},
            "tolerance": {"type": "number", "description": "Comparison tolerance (default 0.01)"},
        },
        "required": ["path"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        from .excel_formula import FormulaError, Grid, evaluate_formula

        path = _resolve_read_path(str(args.get("path", "")), ctx)
        ok, reason = fs_read_ok(str(path))
        if not ok:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"not an Excel workbook: {path.name}")

        def _check() -> dict[str, Any]:
            grid = Grid(path)
            try:
                sheet = str(args.get("sheet") or "").strip() or grid.active
                formula = str(args.get("formula") or "").strip()
                cell = str(args.get("cell") or "").strip()
                if not formula and not cell:
                    raise FormulaError("pass a formula or a cell to check")
                cached = None
                if cell and not formula:
                    from openpyxl import load_workbook

                    fwb = load_workbook(str(path), data_only=False, read_only=True)
                    raw = fwb[sheet][cell].value if sheet in fwb.sheetnames else None
                    fwb.close()
                    if not (isinstance(raw, str) and raw.startswith("=")):
                        raise FormulaError(f"cell {cell} holds no formula (value: {raw!r})")
                    formula = raw
                    cached = grid.get(sheet, *_cell_cr(cell))
                computed = evaluate_formula(formula, grid, sheet)
                out: dict[str, Any] = {
                    "formula": formula, "sheet": sheet, "computed": computed,
                }
                tol = float(args.get("tolerance") or 0.01)
                target = None
                target_src = ""
                if args.get("expected") is not None:
                    target, target_src = float(args["expected"]), "expected"
                elif str(args.get("expect_cell") or "").strip():
                    ec = str(args["expect_cell"]).strip()
                    target = grid.get(sheet, *_cell_cr(ec))
                    target_src = f"cell {ec}"
                elif cached is not None:
                    target, target_src = cached, f"cached value of {cell}"
                if target is not None:
                    try:
                        delta = abs(float(computed) - float(target))
                    except (TypeError, ValueError):
                        raise FormulaError(
                            f"cannot compare {computed!r} to {target!r} numerically"
                        )
                    out.update(target=float(target), target_source=target_src,
                               delta=delta, match=delta <= tol, tolerance=tol)
                elif cell and cached is None:
                    out["note"] = (
                        "no cached value to compare against (the file was not "
                        "saved by Excel) — pass expected= or expect_cell="
                    )
                return out
            finally:
                grid.close()

        try:
            data = await asyncio.to_thread(_check)
        except FormulaError as exc:
            return ToolResult(ok=False, error=str(exc))
        except Exception as exc:  # noqa: BLE001 — real files must not crash the runtime
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        comp = data["computed"]
        comp_s = f"{comp:,.2f}" if isinstance(comp, (int, float)) else repr(comp)
        if "match" in data:
            verdict = (
                f"MATCH (Δ {data['delta']:.4f} ≤ {data['tolerance']})"
                if data["match"]
                else f"MISMATCH — computed {comp_s} vs {data['target']:,.2f} "
                     f"({data['target_source']}, Δ {data['delta']:.4f})"
            )
            out = f"{data['formula']} on {data['sheet']} = {comp_s} → {verdict}"
        else:
            out = f"{data['formula']} on {data['sheet']} = {comp_s}"
            if data.get("note"):
                out += f" ({data['note']})"
        return ToolResult(ok=True, output=out, data=data)


def _cell_cr(ref: str) -> tuple[int, int]:
    """A1 → (col, row) for Grid.get."""
    import re as _re

    m = _re.match(r"^\$?([A-Za-z]{1,3})\$?([0-9]{1,7})$", ref.strip())
    if m is None:
        raise ValueError(f"not a cell reference: {ref!r}")
    from openpyxl.utils import column_index_from_string

    return column_index_from_string(m.group(1).upper()), int(m.group(2))


class ExcelSheetSpecTool(Tool):
    name = "excel_sheet_spec"
    reversibility = Reversibility.READONLY
    returns_untrusted_content = True
    description = (
        "Capture a sheet's REUSABLE STRUCTURE as a spec: every formula cell "
        "(with its cached value when the file was saved by Excel), non-default "
        "formatting (number formats, bold, fills), column widths, merges, and "
        "the account labels in `label_column`. Feed the spec to "
        "excel_apply_spec to reproduce the sheet elsewhere, and save the "
        "workflow as a skill (skill_create) for future reuse."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string", "description": "Sheet name (default: active)"},
            "label_column": {"type": "string", "description": "Account-label column (default A)"},
        },
        "required": ["path"],
    }

    _MAX_SPEC_CELLS = 400

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        path = _resolve_read_path(str(args.get("path", "")), ctx)
        ok, reason = fs_read_ok(str(path))
        if not ok:
            return ToolResult(ok=False, error=f"read denied: {reason}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"not an Excel workbook: {path.name}")

        def _capture() -> dict[str, Any]:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter, column_index_from_string

            fwb = load_workbook(str(path), data_only=False)
            vwb = load_workbook(str(path), data_only=True)
            want = str(args.get("sheet") or "").strip()
            fws = fwb[want] if want and want in fwb.sheetnames else fwb.active
            vws = vwb[fws.title]
            label_col = str(args.get("label_column") or "A").strip().upper()
            li = column_index_from_string(label_col)
            cells: dict[str, dict[str, Any]] = {}
            truncated = False
            for row in fws.iter_rows():
                for c in row:
                    entry: dict[str, Any] = {}
                    if isinstance(c.value, str) and c.value.startswith("="):
                        entry["formula"] = c.value
                        cached = vws.cell(row=c.row, column=c.column).value
                        if isinstance(cached, (int, float)) and not isinstance(cached, bool):
                            entry["value"] = float(cached)
                    if c.number_format and c.number_format != "General":
                        entry["number_format"] = c.number_format
                    if c.font is not None and bool(c.font.bold):
                        entry["bold"] = True
                    fill = c.fill
                    if (
                        fill is not None and fill.patternType == "solid"
                        and getattr(fill.fgColor, "rgb", None)
                        and fill.fgColor.rgb != "00000000"
                    ):
                        entry["fill"] = str(fill.fgColor.rgb)
                    if entry:
                        if len(cells) >= self._MAX_SPEC_CELLS:
                            truncated = True
                            break
                        cells[c.coordinate] = entry
                if truncated:
                    break
            labels: dict[str, str] = {}
            for r in range(1, min(fws.max_row, 2000) + 1):
                v = fws.cell(row=r, column=li).value
                if v is not None and str(v).strip() and not str(v).startswith("="):
                    labels[str(r)] = str(v).strip()
            widths = {
                letter: dim.width
                for letter, dim in fws.column_dimensions.items()
                if dim.width is not None
            }
            merges = [str(m) for m in fws.merged_cells.ranges]
            spec = {
                "sheet": fws.title, "label_column": label_col, "labels": labels,
                "cells": cells, "column_widths": widths, "merges": merges,
                "truncated": truncated,
            }
            fwb.close()
            vwb.close()
            return spec

        try:
            spec = await asyncio.to_thread(_capture)
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        formulas = sum(1 for c in spec["cells"].values() if "formula" in c)
        out = (
            f"spec of {spec['sheet']}: {formulas} formula cell(s), "
            f"{len(spec['cells'])} styled/formula cell(s), "
            f"{len(spec['labels'])} label(s) in column {spec['label_column']}, "
            f"{len(spec['merges'])} merge(s)"
            + (" [cell capture capped at 400]" if spec["truncated"] else "")
        )
        return ToolResult(ok=True, output=out, data=spec)


class ExcelApplySpecTool(Tool):
    name = "excel_apply_spec"
    reversibility = Reversibility.REVERSIBLE  # TX-01: prior bytes captured
    description = (
        "REPRODUCE a captured sheet spec (excel_sheet_spec) onto a workbook in "
        "the workspace: writes the formulas, number formats, bold/fills, "
        "column widths and merges, then VALIDATES by computing every written "
        "formula against the target's real values. With "
        "verify_against_source=true it also compares each computed result to "
        "the spec's recorded values — proving the reproduction matches the "
        "original. Creates the file/sheet when missing. Undoable."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Workspace-relative target workbook"},
            "sheet": {"type": "string", "description": "Target sheet (default: the spec's)"},
            "spec": {"type": "object", "description": "A spec from excel_sheet_spec"},
            "verify_against_source": {
                "type": "boolean",
                "description": "Also compare computed formulas to the spec's recorded values",
            },
            "tolerance": {"type": "number", "description": "Comparison tolerance (default 0.01)"},
        },
        "required": ["path", "spec"],
    }

    async def capture_undo(self, args: dict[str, Any], ctx: ToolContext) -> "dict[str, Any] | None":
        try:
            target = safe_path(ctx.workspace, args["path"])
        except Exception:  # noqa: BLE001
            return None
        if not target.is_file():
            return None
        try:
            prior = target.read_bytes()
        except OSError:
            return None
        return make_file_descriptor(
            ctx.config.home, kind="file_restore", path=args["path"],
            mode="raw", prior_bytes=prior, pre_sha256=sha256_bytes(prior),
        )

    async def revert(self, undo: dict[str, Any], ctx: ToolContext) -> ToolResult:
        return await revert_workspace_file(undo, ctx)

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        from .excel_formula import FormulaError, Grid, evaluate_formula

        try:
            target = safe_path(ctx.workspace, args["path"])
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=str(exc))
        spec = args.get("spec")
        if not isinstance(spec, dict) or not isinstance(spec.get("cells"), dict):
            return ToolResult(ok=False, error="spec must be an excel_sheet_spec object")
        if target.suffix.lower() not in (".xlsx", ".xlsm"):
            return ToolResult(ok=False, error=f"target must be a workbook: {target.name}")

        def _apply() -> dict[str, Any]:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill

            sheet_name = (
                str(args.get("sheet") or "").strip()
                or str(spec.get("sheet") or "").strip() or "Sheet1"
            )
            if target.is_file():
                wb = load_workbook(str(target))
            else:
                wb = Workbook()
                wb.active.title = sheet_name
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            applied = {"formulas": 0, "formats": 0}
            formula_cells: list[str] = []
            for ref, entry in spec["cells"].items():
                if not isinstance(entry, dict):
                    continue
                if isinstance(entry.get("formula"), str):
                    ws[ref] = entry["formula"]
                    applied["formulas"] += 1
                    formula_cells.append(ref)
                styled = False
                if entry.get("number_format"):
                    ws[ref].number_format = str(entry["number_format"])
                    styled = True
                if entry.get("bold"):
                    ws[ref].font = Font(bold=True)
                    styled = True
                if entry.get("fill"):
                    rgb = str(entry["fill"])
                    ws[ref].fill = PatternFill(
                        start_color=rgb, end_color=rgb, fill_type="solid"
                    )
                    styled = True
                if styled:
                    applied["formats"] += 1
            for letter, width in (spec.get("column_widths") or {}).items():
                try:
                    ws.column_dimensions[str(letter)].width = float(width)
                except (TypeError, ValueError):
                    pass
            for rng in spec.get("merges") or []:
                try:
                    ws.merge_cells(str(rng))
                except Exception:  # noqa: BLE001 — a bad merge must not abort the rest
                    pass
            wb.save(str(target))

            # VALIDATE: compute every written formula against the saved target.
            tol = float(args.get("tolerance") or 0.01)
            verify = bool(args.get("verify_against_source"))
            grid = Grid(target)
            checks: list[dict[str, Any]] = []
            failed = 0
            try:
                for ref in formula_cells:
                    entry = spec["cells"][ref]
                    row: dict[str, Any] = {"cell": ref, "formula": entry["formula"]}
                    try:
                        computed = evaluate_formula(entry["formula"], grid, sheet_name)
                        row["computed"] = computed
                        if verify and isinstance(entry.get("value"), (int, float)):
                            delta = abs(float(computed) - float(entry["value"]))
                            row.update(source_value=float(entry["value"]),
                                       delta=delta, match=delta <= tol)
                            if not row["match"]:
                                failed += 1
                    except FormulaError as exc:
                        row["error"] = str(exc)  # honest: couldn't verify this one
                    checks.append(row)
            finally:
                grid.close()
            return {"sheet": sheet_name, "applied": applied,
                    "validation": checks, "failed": failed, "verified": verify}

        try:
            data = await asyncio.to_thread(_apply)
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        rel = str(target.relative_to(Path(ctx.workspace).resolve())).replace("\\", "/")
        head = (
            f"VALIDATION FAILED — {data['failed']} formula(s) do not match the source"
            if data["failed"]
            else "reproduced + validated"
        )
        out = (
            f"{head}: {data['applied']['formulas']} formula(s) + "
            f"{data['applied']['formats']} formatted cell(s) on {data['sheet']} of {rel}"
        )
        return ToolResult(ok=True, output=out, data={"path": rel, **data})


class ExcelAccountsDiffTool(Tool):
    name = "excel_accounts_diff"
    reversibility = Reversibility.READONLY
    returns_untrusted_content = True
    description = (
        "Compare the ACCOUNT LABELS of two sheets (or two versions of a "
        "statement): which accounts were added, removed, or moved to a "
        "different row. Use it when a financial statement changed to see "
        "exactly which label-anchored ranges a working formula must absorb — "
        "then rebuild the formula from the new rows and prove it with "
        "excel_formula_check. Any policy-allowed paths."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path_a": {"type": "string", "description": "The ORIGINAL workbook"},
            "sheet_a": {"type": "string"},
            "path_b": {"type": "string", "description": "The CURRENT workbook (default: path_a)"},
            "sheet_b": {"type": "string"},
            "label_column": {"type": "string", "description": "Account-label column (default A)"},
        },
        "required": ["path_a"],
    }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        pa = _resolve_read_path(str(args.get("path_a", "")), ctx)
        pb = _resolve_read_path(str(args.get("path_b") or args.get("path_a", "")), ctx)
        for p in (pa, pb):
            ok, reason = fs_read_ok(str(p))
            if not ok:
                return ToolResult(ok=False, error=f"read denied: {reason}")
            if p.suffix.lower() not in (".xlsx", ".xlsm"):
                return ToolResult(ok=False, error=f"not an Excel workbook: {p.name}")

        def _labels(path: Path, sheet: "str | None") -> dict[str, int]:
            from openpyxl import load_workbook
            from openpyxl.utils import column_index_from_string

            wb = load_workbook(str(path), data_only=True, read_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            li = column_index_from_string(
                str(args.get("label_column") or "A").strip().upper()
            )
            out: dict[str, int] = {}
            for r, row in enumerate(ws.iter_rows(min_col=li, max_col=li,
                                                 values_only=True), start=1):
                if r > 5000:
                    break
                v = row[0]
                label = str(v).strip() if v is not None else ""
                if label and label not in out:  # first occurrence anchors
                    out[label] = r
            wb.close()
            return out

        try:
            a = await asyncio.to_thread(_labels, pa, args.get("sheet_a"))
            b = await asyncio.to_thread(
                _labels, pb, args.get("sheet_b") or args.get("sheet_a")
            )
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=f"{type(exc).__name__}: {exc}")
        added = sorted((lbl, r) for lbl, r in b.items() if lbl not in a)
        removed = sorted((lbl, r) for lbl, r in a.items() if lbl not in b)
        moved = sorted(
            (lbl, a[lbl], b[lbl]) for lbl in a if lbl in b and a[lbl] != b[lbl]
        )
        lines = [
            f"accounts: {len(a)} before, {len(b)} now — "
            f"{len(added)} added, {len(removed)} removed, {len(moved)} moved"
        ]
        lines += [f"+ {lbl} (row {r})" for lbl, r in added[:20]]
        lines += [f"- {lbl} (was row {r})" for lbl, r in removed[:20]]
        lines += [f"~ {lbl}: row {r1} → {r2}" for lbl, r1, r2 in moved[:20]]
        if added or removed or moved:
            lines.append(
                "Update label-anchored ranges to the NEW rows, then prove the "
                "updated formula with excel_formula_check before writing it."
            )
        return ToolResult(ok=True, output="\n".join(lines), data={
            "added": [{"label": lbl, "row": r} for lbl, r in added],
            "removed": [{"label": lbl, "row": r} for lbl, r in removed],
            "moved": [{"label": lbl, "from": r1, "to": r2} for lbl, r1, r2 in moved],
            "labels_now": b,
        })


def excel_tools() -> list[Tool]:
    return [
        ExcelReadTool(), ExcelEditTool(), ExcelProfileTool(), ExcelQueryTool(),
        ExcelFormulaCheckTool(), ExcelSheetSpecTool(), ExcelApplySpecTool(),
        ExcelAccountsDiffTool(),
    ]
