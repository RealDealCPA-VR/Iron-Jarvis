"""Excel formula engine (v1.90.0): parse + evaluate + validate IN the engine.

openpyxl reads formulas but never computes them, and a formula the agent just
wrote has no cached value until real Excel opens the file — so "write a formula
and CHECK it against the sheet" needs our own evaluator. This is a deliberate
SUBSET built for financial-statement work: cell refs (absolute/relative,
sheet-qualified, quoted sheet names), ranges, arithmetic (+ - * / ^ %),
comparisons, and the functions that cover totals/subtotals/rollups:

    SUM AVERAGE MIN MAX COUNT COUNTA ABS ROUND IF SUBTOTAL(9|1)

Anything outside the subset raises :class:`FormulaError` naming the function —
an honest "I can't verify that one" beats a silently wrong number. Excel
semantics where they matter: blank cells are 0 in arithmetic, SUM/AVERAGE skip
text, division by zero is an error.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

#: Range-iteration cap — a whole-column ref (A:A is unsupported anyway) or a
#: giant range must not stall the daemon.
_MAX_RANGE_CELLS = 200_000


class FormulaError(ValueError):
    """A formula this engine cannot parse/evaluate — message is user-facing."""


_TOKEN_RX = re.compile(
    r"""
    (?P<ws>\s+)
  | (?P<sheet>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)
  | (?P<cell>\$?[A-Za-z]{1,3}\$?[0-9]{1,7})
  | (?P<number>[0-9]*\.?[0-9]+(?:[eE][+-]?[0-9]+)?)
  | (?P<string>"(?:[^"]|"")*")
  | (?P<ident>[A-Za-z_][A-Za-z0-9_.]*)
  | (?P<op><=|>=|<>|[=<>+\-*/^%(),:])
    """,
    re.VERBOSE,
)

_CELL_RX = re.compile(r"^\$?([A-Za-z]{1,3})\$?([0-9]{1,7})$")


def _col_num(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


class Grid:
    """Cell values of one workbook (``data_only=True``): get(sheet, col, row)."""

    def __init__(self, path: "str | Path") -> None:
        from openpyxl import load_workbook

        self._wb = load_workbook(str(path), data_only=True, read_only=False)
        self.sheetnames = list(self._wb.sheetnames)
        self.active = self._wb.active.title

    def get(self, sheet: str, col: int, row: int) -> Any:
        try:
            ws = self._wb[sheet]
        except KeyError:
            raise FormulaError(f"no sheet named {sheet!r}")
        return ws.cell(row=row, column=col).value

    def close(self) -> None:
        try:
            self._wb.close()
        except Exception:  # noqa: BLE001
            pass


def _tokenize(src: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    pos = 0
    while pos < len(src):
        m = _TOKEN_RX.match(src, pos)
        if m is None:
            raise FormulaError(f"unexpected character {src[pos]!r} in formula")
        pos = m.end()
        kind = m.lastgroup or ""
        if kind == "ws":
            continue
        out.append((kind, m.group()))
    return out


class _Parser:
    """Recursive descent with Excel-ish precedence:
    comparison < add < mul < unary < power < percent < primary."""

    def __init__(self, tokens: list[tuple[str, str]], grid: Grid, sheet: str) -> None:
        self.toks = tokens
        self.i = 0
        self.grid = grid
        self.sheet = sheet

    # -- token helpers -----------------------------------------------------
    def _peek(self) -> tuple[str, str]:
        return self.toks[self.i] if self.i < len(self.toks) else ("eof", "")

    def _take(self) -> tuple[str, str]:
        t = self._peek()
        self.i += 1
        return t

    def _expect(self, value: str) -> None:
        kind, val = self._take()
        if val != value:
            raise FormulaError(f"expected {value!r}, found {val or 'end of formula'!r}")

    # -- grammar -----------------------------------------------------------
    def parse(self) -> Any:
        v = self._comparison()
        if self._peek()[0] != "eof":
            raise FormulaError(f"unexpected {self._peek()[1]!r} after expression")
        return v

    def _comparison(self) -> Any:
        left = self._additive()
        while self._peek()[1] in ("=", "<>", "<", ">", "<=", ">="):
            op = self._take()[1]
            right = self._additive()
            ln, rn = _num_or_str(left), _num_or_str(right)
            if op == "=":
                left = ln == rn
            elif op == "<>":
                left = ln != rn
            else:
                a, b = _as_num(left), _as_num(right)
                left = {"<": a < b, ">": a > b, "<=": a <= b, ">=": a >= b}[op]
        return left

    def _additive(self) -> Any:
        left = self._multiplicative()
        while self._peek()[1] in ("+", "-"):
            op = self._take()[1]
            right = self._multiplicative()
            a, b = _as_num(left), _as_num(right)
            left = a + b if op == "+" else a - b
        return left

    def _multiplicative(self) -> Any:
        left = self._unary()
        while self._peek()[1] in ("*", "/"):
            op = self._take()[1]
            right = self._unary()
            a, b = _as_num(left), _as_num(right)
            if op == "/":
                if b == 0:
                    raise FormulaError("division by zero")
                left = a / b
            else:
                left = a * b
        return left

    def _unary(self) -> Any:
        if self._peek()[1] in ("+", "-"):
            op = self._take()[1]
            v = _as_num(self._unary())
            return v if op == "+" else -v
        return self._power()

    def _power(self) -> Any:
        left = self._percent()
        if self._peek()[1] == "^":
            self._take()
            right = self._unary()  # right-associative
            return _as_num(left) ** _as_num(right)
        return left

    def _percent(self) -> Any:
        v = self._primary()
        while self._peek()[1] == "%":
            self._take()
            v = _as_num(v) / 100.0
        return v

    def _primary(self) -> Any:
        kind, val = self._take()
        if kind == "number":
            return float(val)
        if kind == "string":
            return val[1:-1].replace('""', '"')
        if val == "(":
            v = self._comparison()
            self._expect(")")
            return v
        if kind == "sheet":
            name = val[:-1]
            if name.startswith("'"):
                name = name[1:-1]
            ck, cv = self._take()
            if ck != "cell":
                raise FormulaError(f"expected a cell after {name!r}!")
            return self._ref(name, cv)
        if kind == "cell":
            return self._ref(self.sheet, val)
        if kind == "ident":
            if self._peek()[1] != "(":
                raise FormulaError(
                    f"named ranges are not supported (saw {val!r}) — use cell refs"
                )
            self._take()  # (
            args: list[Any] = []
            if self._peek()[1] != ")":
                args.append(self._comparison())
                while self._peek()[1] == ",":
                    self._take()
                    args.append(self._comparison())
            self._expect(")")
            return _call(val.upper(), args)
        raise FormulaError(f"unexpected {val or 'end of formula'!r}")

    def _ref(self, sheet: str, first: str) -> Any:
        """A cell value, or a _Range when a ':' follows."""
        if self._peek()[1] == ":":
            self._take()
            nk, nv = self._take()
            if nk == "sheet":  # tolerate Sheet1!A1:Sheet1!B2 (same sheet only)
                nk, nv = self._take()
            if nk != "cell":
                raise FormulaError("expected a cell after ':' in a range")
            return _Range(self.grid, sheet, first, nv)
        m = _CELL_RX.match(first)
        assert m is not None  # the tokenizer only produces valid cells
        return self.grid.get(sheet, _col_num(m.group(1)), int(m.group(2)))


class _Range:
    """A rectangular range — only meaningful inside range-taking functions."""

    def __init__(self, grid: Grid, sheet: str, a: str, b: str) -> None:
        ma, mb = _CELL_RX.match(a), _CELL_RX.match(b)
        assert ma is not None and mb is not None
        c1, r1 = _col_num(ma.group(1)), int(ma.group(2))
        c2, r2 = _col_num(mb.group(1)), int(mb.group(2))
        self.grid, self.sheet = grid, sheet
        self.c1, self.c2 = min(c1, c2), max(c1, c2)
        self.r1, self.r2 = min(r1, r2), max(r1, r2)
        if (self.c2 - self.c1 + 1) * (self.r2 - self.r1 + 1) > _MAX_RANGE_CELLS:
            raise FormulaError("range too large to evaluate")

    def values(self) -> list[Any]:
        out = []
        for r in range(self.r1, self.r2 + 1):
            for c in range(self.c1, self.c2 + 1):
                out.append(self.grid.get(self.sheet, c, r))
        return out


def _as_num(v: Any) -> float:
    if isinstance(v, _Range):
        raise FormulaError("a range can only be used inside SUM/AVERAGE/… functions")
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        raise FormulaError(f"cannot use text {str(v)[:40]!r} as a number")


def _num_or_str(v: Any) -> Any:
    if isinstance(v, _Range):
        raise FormulaError("a range can only be used inside SUM/AVERAGE/… functions")
    if isinstance(v, str):
        return v.strip().lower()
    return _as_num(v)


def _numeric_items(args: list[Any]) -> list[float]:
    """Flatten args (values + ranges) to the NUMERIC items, Excel-style: range
    cells that hold text/blank are skipped; literal args must be numeric."""
    out: list[float] = []
    for a in args:
        if isinstance(a, _Range):
            for v in a.values():
                if isinstance(v, bool) or v is None or isinstance(v, str):
                    continue
                if isinstance(v, (int, float)):
                    out.append(float(v))
        else:
            out.append(_as_num(a))
    return out


def _call(name: str, args: list[Any]) -> Any:
    if name == "SUM":
        return sum(_numeric_items(args))
    if name == "AVERAGE":
        items = _numeric_items(args)
        if not items:
            raise FormulaError("AVERAGE of no numeric values")
        return sum(items) / len(items)
    if name == "MIN":
        items = _numeric_items(args)
        return min(items) if items else 0.0
    if name == "MAX":
        items = _numeric_items(args)
        return max(items) if items else 0.0
    if name == "COUNT":
        return float(len(_numeric_items(args)))
    if name == "COUNTA":
        n = 0
        for a in args:
            if isinstance(a, _Range):
                n += sum(1 for v in a.values() if v is not None and v != "")
            elif a is not None and a != "":
                n += 1
        return float(n)
    if name == "ABS":
        if len(args) != 1:
            raise FormulaError("ABS takes exactly one argument")
        return abs(_as_num(args[0]))
    if name == "ROUND":
        if len(args) not in (1, 2):
            raise FormulaError("ROUND takes (value, digits)")
        digits = int(_as_num(args[1])) if len(args) == 2 else 0
        return round(_as_num(args[0]), digits)
    if name == "IF":
        if len(args) not in (2, 3):
            raise FormulaError("IF takes (condition, then, else)")
        cond = args[0]
        truthy = bool(cond) if isinstance(cond, bool) else _as_num(cond) != 0
        return args[1] if truthy else (args[2] if len(args) == 3 else False)
    if name == "SUBTOTAL":
        if len(args) < 2:
            raise FormulaError("SUBTOTAL takes (function_num, range…)")
        fn = int(_as_num(args[0]))
        if fn in (9, 109):
            return sum(_numeric_items(args[1:]))
        if fn in (1, 101):
            items = _numeric_items(args[1:])
            if not items:
                raise FormulaError("SUBTOTAL average of no numeric values")
            return sum(items) / len(items)
        raise FormulaError(f"SUBTOTAL({fn}, …) is not supported (9=sum, 1=average)")
    raise FormulaError(
        f"function {name} is not supported by the checker — supported: SUM,"
        " AVERAGE, MIN, MAX, COUNT, COUNTA, ABS, ROUND, IF, SUBTOTAL"
    )


def evaluate_formula(formula: str, grid: Grid, sheet: str) -> Any:
    """Evaluate *formula* (with or without a leading '=') against *grid*,
    resolving bare cell refs on *sheet*. Raises :class:`FormulaError` with an
    honest message for anything outside the supported subset."""
    src = (formula or "").strip()
    if src.startswith("="):
        src = src[1:]
    if not src:
        raise FormulaError("empty formula")
    tokens = _tokenize(src)
    result = _Parser(tokens, grid, sheet).parse()
    if isinstance(result, _Range):
        raise FormulaError("a bare range is not a value — wrap it in SUM(...)")
    if isinstance(result, bool):
        return result
    return result
