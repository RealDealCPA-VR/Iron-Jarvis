"""Document routes: uploads, living documents, read/write, enhance.

Moved verbatim from daemon/app.py's create_app; closure-local state is
reached through ``d`` (see the deps object built in create_app).
"""

from __future__ import annotations

from fastapi import FastAPI, HTTPException
from sqlmodel import select
from pathlib import Path
from typing import Any

from .. import app as _app
from ..schemas import (
    DocEnhanceBody,
    DocumentOpenBody,
    DocWriteBody,
    LiveDocCreate,
    RedactApplyBody,
    RedactScanBody,
    SaveCopyBody,
    UploadBody,
)
from ...core.db import session_scope
from ...core.fs_policy import fs_read_ok

#: Suffixes the preview panel renders as pixels (v1.166.0). The preview
#: endpoint returns a POINTER only — the client loads the bytes via
#: /documents/file, which serves these inline so an <img>/iframe can show them.
_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
#: What may render inline in the browser; everything else downloads
#: (xlsx/docx can't render in an iframe anyway).
_INLINE_SUFFIXES = {".pdf", *_IMAGE_SUFFIXES}

#: Suffix → the app a native open will land in (user-facing button label).
_APP_LABEL = {
    ".docx": "Word", ".doc": "Word",
    ".xlsx": "Excel", ".xlsm": "Excel", ".csv": "Excel",
    ".pptx": "PowerPoint",
    ".pdf": "your PDF viewer",
    ".html": "your browser",
}


def _gated_source(raw: str, base: Path | None = None) -> Path:
    """Resolve a caller-supplied SOURCE path first, then gate the RESOLVED path.

    The order is the whole point. ``fs_read_ok``/``is_protected_path`` resolve
    what they are handed, but ``Path.resolve()`` does NOT expand ``~`` and
    cannot know about a join that happens afterwards — so gating the RAW
    request string and then opening the ``expanduser()``-ed / ``base``-joined
    path checks one file and reads another. Both forms were live bypasses:
    ``~/.ironjarvis/ironjarvis.db`` gated as a literal ``~`` folder under the
    cwd, and a relative ``../secrets/<file>`` gated against the cwd, while the
    read landed inside a registered protected root (the Fernet key material,
    the SQLite DB holding inline undo pre-images of the user's real files).

    *base* is the directory a RELATIVE path is joined to; ``None`` means the
    caller must pass an absolute path, as the sibling creative.py file routes
    and ``_preview_path`` above already require.
    """
    text = (raw or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="a path is required")
    src = Path(text).expanduser()
    if not src.is_absolute():
        if base is None:
            raise HTTPException(status_code=400, detail="an absolute path is required")
        src = base / src
    src = src.resolve()
    ok, reason = fs_read_ok(str(src))
    if not ok:
        raise HTTPException(status_code=403, detail=reason)
    return src


def _open_native(path: str) -> None:
    """Launch *path* with the OS-associated application (Word/Excel/…).
    Module-level so tests monkeypatch it instead of really launching apps."""
    import os
    import subprocess
    import sys

    if sys.platform == "win32":
        os.startfile(path)  # noqa: S606 — explicit, user-initiated open
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


def register(app: FastAPI, d) -> None:
    """Attach these routes to *app*; ``d`` is the create_app deps object."""
    @app.post("/documents/upload")
    def documents_upload(body: UploadBody) -> dict[str, Any]:
        """Accept a base64 file and store it under <home>/uploads (no multipart dep)."""
        import base64
        import re

        # Cap the decoded size so a giant upload can't OOM-kill the whole daemon
        # (which would take down every session/terminal with it). 4/3 accounts for
        # base64 expansion; reject BEFORE decoding so we never buffer the bytes.
        approx_bytes = (len(body.content_b64) * 3) // 4
        if approx_bytes > _app._MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"upload too large (~{approx_bytes // (1024 * 1024)} MB); "
                    f"limit is {_app._MAX_UPLOAD_BYTES // (1024 * 1024)} MB"
                ),
            )
        name = re.sub(r"[^A-Za-z0-9._-]", "_", body.filename).strip("._") or "upload"
        uploads = d.platform.config.home / "uploads"
        uploads.mkdir(parents=True, exist_ok=True)
        target = uploads / name
        try:
            data = base64.b64decode(body.content_b64, validate=False)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"invalid base64: {exc}")
        target.write_bytes(data)
        return {"path": str(target), "name": name, "bytes": len(data)}

    # ------------------------------------------------------------------ #
    # Preview + native open (v1.89.0) — the chat's embedded document panel.
    # Any policy-allowed ABSOLUTE path works: local disk, a network share,
    # or a tailnet folder (the daemon reads it like any other file).
    # ------------------------------------------------------------------ #
    def _preview_path(raw: str) -> Path:
        from ...core.fs_policy import is_protected_path

        p = Path((raw or "").strip())
        if not raw or not p.is_absolute():
            raise HTTPException(status_code=400, detail="an absolute path is required")
        ok, reason = fs_read_ok(str(p))
        if not ok or is_protected_path(str(p)):
            raise HTTPException(status_code=403, detail=reason or "path not allowed")
        if not p.is_file():
            raise HTTPException(status_code=404, detail=f"no such file: {p}")
        return p

    @app.get("/documents/preview")
    def document_preview(path: str, sheet: str = "") -> dict[str, Any]:
        """Structured preview of ONE document: spreadsheets as sheet tabs +
        rows (engine-read, capped), PDFs as an embed pointer (the client
        iframes /documents/file), everything else as extracted text."""
        p = _preview_path(path)
        suffix = p.suffix.lower()
        base: dict[str, Any] = {"name": p.name, "path": str(p), "suffix": suffix}
        if suffix in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            try:
                wb = load_workbook(str(p), data_only=True, read_only=True)
                names = list(wb.sheetnames)
                ws = wb[sheet] if sheet and sheet in names else wb.active
                rows: list[list[str]] = []
                truncated = False
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 80:
                        truncated = True
                        break
                    # Column and cell clips are TRUNCATION too (v1.167.0): a
                    # 40-column ledger shown as a complete-looking 30-column
                    # table — and the Changes diff then reporting "no changes"
                    # for an edit in column AK — is the silent-truncation class.
                    if len(row) > 30:
                        truncated = True
                    if any(
                        v is not None and len(str(v)) > 80 for v in row[:30]
                    ):
                        truncated = True
                    rows.append(["" if v is None else str(v)[:80] for v in row[:30]])
                title = ws.title
                # Truncation honesty (v1.166.0): the REAL extent, so the client
                # can say "first 80 of N rows" instead of implying completeness.
                total_rows = ws.max_row
                total_cols = ws.max_column
                wb.close()
            except Exception as exc:  # noqa: BLE001
                raise HTTPException(
                    status_code=422, detail=f"could not read workbook: {exc}"
                )
            out = {**base, "kind": "sheet", "sheets": names, "sheet": title,
                   "rows": rows, "truncated": truncated}
            if total_rows is not None:  # read-only mode may lack dimensions
                out["total_rows"] = int(total_rows)
            if total_cols is not None:
                out["total_cols"] = int(total_cols)
            return out
        if suffix == ".pdf":
            return {**base, "kind": "pdf"}
        if suffix in _IMAGE_SUFFIXES:
            # A pointer only — no bytes in the JSON. The client loads the
            # pixels via /documents/file (served inline for image suffixes).
            return {**base, "kind": "image"}
        if suffix in (".csv", ".tsv"):
            import csv as _csv

            delim = "\t" if suffix == ".tsv" else ","
            rows = []
            truncated = False
            total_rows = 0
            total_cols = 0
            try:
                with p.open("r", encoding="utf-8", errors="replace", newline="") as fh:
                    for row in _csv.reader(fh, delimiter=delim):
                        total_rows += 1  # real count — full iteration is cheap
                        total_cols = max(total_cols, len(row))
                        if len(rows) >= 80:
                            truncated = True
                            continue
                        if len(row) > 30:
                            truncated = True
                        if any(len(str(v)) > 80 for v in row[:30]):
                            truncated = True  # cell clip is truncation too
                        rows.append([str(v)[:80] for v in row[:30]])
            except (OSError, _csv.Error) as exc:
                # _csv.Error covers parser failures such as a field over the
                # process-wide csv.field_size_limit (~128KB) — a realistic CSV
                # export with a blob cell. Do NOT raise the limit: it is
                # process-global state shared with every other csv consumer.
                raise HTTPException(status_code=422, detail=f"could not read: {exc}")
            return {**base, "kind": "sheet", "sheets": ["CSV"], "sheet": "CSV",
                    "rows": rows, "truncated": truncated,
                    "total_rows": total_rows, "total_cols": total_cols}
        if suffix == ".docx":
            # WORD-FAITHFUL preview: semantic docx→HTML (headings, bold/italic,
            # lists, real tables) rendered by the client on a white page in a
            # SANDBOXED frame — extracted text loses exactly the formatting the
            # user is reviewing. Degrades to structured markdown, then text.
            try:
                import mammoth

                with p.open("rb") as fh:
                    html = mammoth.convert_to_html(fh).value or ""
                if html.strip():
                    return {**base, "kind": "html", "html": html[:400_000],
                            "truncated": len(html) > 400_000,
                            "total_chars": len(html)}
            except Exception:  # noqa: BLE001 — fall through to markdown/text
                pass
            try:
                from ...documents.pdf_markdown import document_to_markdown

                md = document_to_markdown(p)
                if md.strip():
                    return {**base, "kind": "markdown", "content": md[:40_000],
                            "truncated": len(md) > 40_000,
                            "total_chars": len(md)}
            except Exception:  # noqa: BLE001
                pass
        from ...documents.readers import extract_text

        try:
            text = extract_text(p)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=422, detail=f"could not read: {exc}")
        kind = "markdown" if suffix in (".md", ".markdown") else "text"
        return {**base, "kind": kind, "content": text[:20_000],
                "truncated": len(text) > 20_000, "total_chars": len(text)}

    @app.get("/documents/file")
    def document_file(path: str, download: bool = False):
        """Raw file bytes (auth rides the header or ?token= like other embeds)
        — powers the preview panel's PDF iframe and image tags. PDFs and
        images serve INLINE so the browser renders them; ``?download=1``
        forces a save-as, and everything else is attachment always (an xlsx
        can't render in an iframe anyway)."""
        import mimetypes

        from fastapi.responses import FileResponse

        p = _preview_path(path)
        suffix = p.suffix.lower()
        if suffix == ".pdf":
            media = "application/pdf"
        elif suffix in _IMAGE_SUFFIXES:
            guessed, _enc = mimetypes.guess_type(p.name)
            media = guessed or "application/octet-stream"
        else:
            media = "application/octet-stream"
        disposition = (
            "inline" if not download and suffix in _INLINE_SUFFIXES
            else "attachment"
        )
        return FileResponse(str(p), media_type=media, filename=p.name,
                            content_disposition_type=disposition)

    @app.post("/documents/open")
    def document_open(body: DocumentOpenBody) -> dict[str, Any]:
        """Open a document with its OS-associated app (Word/Excel/…) — an
        explicit, user-initiated action from the preview panel's button."""
        p = _preview_path(body.path)
        try:
            _open_native(str(p))
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=422, detail=f"could not open: {exc}")
        return {"ok": True, "app": _APP_LABEL.get(p.suffix.lower(), "the default app")}

    @app.get("/documents/live")
    def list_livedocs() -> dict[str, Any]:
        from ...core.models import LiveDocRecord

        with session_scope(d.platform.engine) as db:
            rows = list(db.exec(select(LiveDocRecord)))
        rows.sort(key=lambda r: r.created_at, reverse=True)
        return {"docs": [r.model_dump() for r in rows]}

    @app.post("/documents/live")
    async def create_livedoc(body: LiveDocCreate) -> dict[str, Any]:
        from ...core.models import LiveDocRecord

        name = (body.name or "").strip()
        if not name or not (body.prompt or "").strip():
            raise HTTPException(status_code=400, detail="name and prompt are required")
        if body.format not in ("md", "html", "docx", "pdf"):
            raise HTTPException(status_code=400, detail="format must be md|html|docx|pdf")
        rec = LiveDocRecord(name=name, prompt=body.prompt.strip(), format=body.format,
                            provider=body.provider, model=body.model)
        with session_scope(d.platform.engine) as db:
            db.add(rec)
            db.commit()
            db.refresh(rec)
        # Optional auto-refresh: an event-kind schedule the lifespan handler
        # listens for. Manual-only docs simply skip this.
        if body.cron or body.interval_seconds:
            sched_name = f"livedoc_{rec.id}"
            try:
                d.platform.scheduler.add_task(
                    sched_name, body.cron,
                    interval_seconds=body.interval_seconds,
                    kind="event",
                    payload={"type": "livedoc.regenerate", "livedoc_id": rec.id},
                )
                with session_scope(d.platform.engine) as db:
                    row = db.get(LiveDocRecord, rec.id)
                    row.schedule_name = sched_name
                    db.add(row)
                    db.commit()
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f"bad schedule: {exc}")
        # First generation now, so the doc exists immediately.
        result = await d._regenerate_livedoc(rec.id)
        return {**result, "name": name, "schedule": bool(body.cron or body.interval_seconds)}

    @app.post("/documents/live/{doc_id}/regenerate")
    async def regenerate_livedoc_ep(doc_id: str) -> dict[str, Any]:
        return await d._regenerate_livedoc(doc_id)

    @app.delete("/documents/live/{doc_id}")
    def delete_livedoc(doc_id: str) -> dict[str, Any]:
        """Remove the living doc + its schedule from the APP. The generated
        file stays on disk (never delete the user's files)."""
        from ...core.models import LiveDocRecord

        with session_scope(d.platform.engine) as db:
            row = db.get(LiveDocRecord, doc_id)
            if row is None:
                raise HTTPException(status_code=404, detail="no such living document")
            sched = row.schedule_name
            db.delete(row)
            db.commit()
        if sched:
            try:
                d.platform.scheduler.remove(sched)
            except Exception:  # noqa: BLE001 — schedule may already be gone
                pass
        return {"deleted": doc_id, "files_touched": 0}

    @app.post("/documents/enhance")
    async def enhance_document(body: DocEnhanceBody) -> dict[str, Any]:
        """Suggest a better filename + polished content BEFORE creating —
        returned for review; nothing is written until the user confirms."""
        import json as _json

        from ...providers.adapters.base import LLMMessage

        if not (body.content or "").strip() and not (body.filename or "").strip():
            raise HTTPException(status_code=400, detail="nothing to enhance")
        provider = body.provider or d.platform.config.default_provider
        model = body.model or d.platform.config.default_model
        try:
            adapter = d.platform.providers.get(provider, model)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"provider unavailable: {exc}")
        system = (
            "You polish document drafts. Respond with ONLY JSON: "
            '{"filename": "improved-name.ext (keep/choose a sensible extension)", '
            '"content": "the improved document content (markdown allowed)", '
            '"notes": "1-3 short bullets on what you changed and why"}. '
            "Improve clarity/structure/professional tone; NEVER invent facts or "
            "figures; keep the user's meaning."
        )
        user = f"Filename: {body.filename or '(none)'}\n\nContent:\n{(body.content or '')[:10000]}"
        resp, _p, _m = await d._one_shot_complete(
            provider, adapter, system=system,
            messages=[LLMMessage(role="user", content=user)],
        )
        text = resp.text or ""
        start, depth, obj = text.find("{"), 0, ""
        if start >= 0:
            for i in range(start, len(text)):
                depth += (text[i] == "{") - (text[i] == "}")
                if depth == 0:
                    obj = text[start:i + 1]
                    break
        try:
            out = _json.loads(obj)
        except Exception:  # noqa: BLE001
            raise HTTPException(status_code=422, detail="no valid suggestion — try again")
        return {
            "filename": str(out.get("filename") or body.filename),
            "content": str(out.get("content") or body.content),
            "notes": str(out.get("notes") or ""),
        }

    @app.get("/documents/read")
    async def documents_read(path: str) -> dict[str, Any]:
        import asyncio as _asyncio

        from ...documents import extract_text
        from ...documents.ocr import looks_scanned_pdf, ocr_pdf

        ok, reason = fs_read_ok(path)
        if not ok:
            raise HTTPException(status_code=403, detail=reason)
        try:
            text = await _asyncio.to_thread(extract_text, path)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"cannot read: {exc}")
        # SCANNED-PDF fallback: an image-only PDF (death certificate, signed
        # form) has no text layer — recover it via vision OCR instead of
        # returning empty silence. The note always says how the text came to
        # be (or exactly why it couldn't).
        note = ""
        p = Path(path)
        if looks_scanned_pdf(p, text):
            try:
                ocr_text, note = await ocr_pdf(p, d.platform.router)
                if ocr_text:
                    text = ocr_text
            except Exception as exc:  # noqa: BLE001 — OCR failure ≠ read failure
                note = f"scanned PDF — OCR fallback failed ({type(exc).__name__}: {exc})"
        return {"path": path, "text": text[:20000], "note": note}

    # ------------------------------------------------------- save a copy to… ---
    # REPORTED: "the agent should give me options with buttons as to where to
    # store the file."
    #
    # Chat runs its tools inside a confined workspace (the uploads scratch dir,
    # or the grounded project folder), so anything it produces lands THERE by
    # construction — correct for confinement, useless as a place to find a
    # finished document. These two routes let the preview panel offer real
    # destinations instead of leaving the user to hunt for the output.

    def _places() -> list[dict[str, str]]:
        home = Path.home()
        out: list[dict[str, str]] = []
        for key, label, p in (
            ("desktop", "Desktop", home / "Desktop"),
            ("documents", "Documents", home / "Documents"),
            ("downloads", "Downloads", home / "Downloads"),
        ):
            if p.is_dir():
                out.append({"key": key, "label": label, "path": str(p)})
        return out

    @app.get("/documents/places")
    def documents_places() -> dict[str, Any]:
        """Common save destinations that actually exist on this machine."""
        return {"places": _places()}

    @app.post("/documents/save-copy")
    def documents_save_copy(body: SaveCopyBody) -> dict[str, Any]:
        import shutil

        from ...core.fs_policy import is_protected_path

        # Resolve BEFORE gating — see _gated_source. The old order checked the
        # raw string and copied the expanduser()'d file.
        src = _gated_source(body.source)
        if not src.is_file():
            raise HTTPException(status_code=404, detail=f"not a file: {body.source}")
        dest_dir = Path(body.dest_dir).expanduser()
        if not dest_dir.is_absolute():
            raise HTTPException(status_code=400, detail="dest_dir must be absolute")
        if is_protected_path(dest_dir):
            raise HTTPException(status_code=403, detail="destination is a protected path")
        if not dest_dir.is_dir():
            raise HTTPException(status_code=404, detail=f"no such folder: {body.dest_dir}")
        target = dest_dir / (body.name.strip() or src.name)
        if target.resolve() == src.resolve():
            raise HTTPException(status_code=400, detail="that is where the file already is")
        if target.exists() and not body.overwrite:
            raise HTTPException(
                status_code=409, detail=f"{target.name} is already in that folder"
            )
        try:
            shutil.copy2(src, target)
        except OSError as exc:
            raise HTTPException(status_code=400, detail=f"could not save: {exc}")
        return {"path": str(target), "name": target.name, "folder": str(dest_dir)}

    # ---------------------------------------------------------- PII redaction ---
    # REPORTED: the redaction tool "didn't show me which items it recognized as
    # PII for my approval" and "routed to a folder without asking me where to
    # put this".
    #
    # Both were true, and both came from the same place: redaction was only ever
    # reachable through an AGENT. The confirm-first contract lived in a tool
    # description and a skill playbook — advice a model may or may not follow —
    # and the pii-redaction skill in fact jumped straight to redact_pii, so the
    # approval list was never shown and the destination defaulted silently.
    #
    # These two routes make the flow a real, deterministic UI: scan returns the
    # candidates, apply redacts EXACTLY the confirmed values to a destination the
    # user picked. No model in the loop, so the step cannot be skipped.

    def _redact_source(path: str) -> Path:
        # Resolve (expanduser + documents-join) BEFORE gating — see
        # _gated_source. Checking the raw string let '~/…' and '../secrets/…'
        # name a protected store the gate never saw.
        src = _gated_source(path, d.platform.config.home / "documents")
        if not src.is_file():
            raise HTTPException(status_code=404, detail=f"not a file: {path}")
        return src

    @app.post("/documents/redact/scan")
    async def documents_redact_scan(body: RedactScanBody) -> dict[str, Any]:
        import asyncio as _asyncio

        from ...documents.redact import ALL_CATEGORIES, scan_document

        src = _redact_source(body.path)
        cats = [str(c).strip().lower() for c in body.categories if str(c).strip()]
        unknown = [c for c in cats if c not in ALL_CATEGORIES]
        if unknown:
            raise HTTPException(
                status_code=400,
                detail=f"unknown categories: {', '.join(unknown)} — "
                       f"valid: {', '.join(sorted(ALL_CATEGORIES))}",
            )
        try:
            findings = await _asyncio.to_thread(
                scan_document,
                src,
                extra_terms=[t for t in body.extra_terms if t.strip()],
                categories=set(cats) or None,
            )
        except Exception as exc:  # noqa: BLE001 — a bad file must not 500
            raise HTTPException(status_code=400, detail=f"cannot scan: {exc}")
        # The default destination is offered UP FRONT so the user can see and
        # change where the output will land before approving anything.
        default_out = src.with_name(f"{src.stem}.redacted{src.suffix}")
        return {
            "source": str(src),
            "name": src.name,
            "findings": findings,
            "default_output_path": str(default_out),
            "suffix": src.suffix.lower(),
        }

    @app.post("/documents/redact/apply")
    async def documents_redact_apply(body: RedactApplyBody) -> dict[str, Any]:
        import asyncio as _asyncio

        from ...core.fs_policy import is_protected_path
        from ...documents.redact import STYLES, redact_file

        src = _redact_source(body.path)
        style = (body.style or "black").strip().lower()
        if style not in STYLES:
            raise HTTPException(
                status_code=400, detail=f"unknown style {style!r} — use {', '.join(STYLES)}"
            )
        terms = [t for t in body.terms if t and t.strip()]
        if not terms:
            # An empty list would fall through to auto-detection in the engine
            # and redact things the user never ticked. Refuse instead.
            raise HTTPException(
                status_code=400,
                detail="no items confirmed — tick at least one finding to redact",
            )
        target = (
            Path(body.output_path).expanduser()
            if body.output_path.strip()
            else src.with_name(f"{src.stem}.redacted{src.suffix}")
        )
        if not target.is_absolute():
            raise HTTPException(status_code=400, detail="output_path must be absolute")
        if is_protected_path(target):
            raise HTTPException(status_code=403, detail="destination is a protected path")
        if target.resolve() == src.resolve():
            raise HTTPException(
                status_code=400,
                detail="destination must differ from the source — the original is never overwritten",
            )
        if target.exists() and not body.overwrite:
            raise HTTPException(
                status_code=409, detail=f"{target.name} already exists at that location"
            )
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            counts, note = await _asyncio.to_thread(
                redact_file, src, target, style=style, only_terms=terms
            )
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"cannot redact: {exc}")
        return {
            "path": str(target),
            "name": target.name,
            "source": str(src),
            "style": style,
            "counts": counts,
            "total": sum(counts.values()),
            "note": note,
        }

    @app.post("/documents/write")
    def documents_write(body: DocWriteBody) -> dict[str, Any]:
        from ...documents import write_document

        base = (d.platform.config.home / "documents").resolve()
        target = (base / body.path).resolve()
        if target != base and not target.is_relative_to(base):
            raise HTTPException(status_code=400, detail="path escapes documents dir")
        out = write_document(target, body.content, kind=body.kind)
        return {
            "path": str(out.relative_to(base)).replace("\\", "/"),
            "bytes": out.stat().st_size,
        }
