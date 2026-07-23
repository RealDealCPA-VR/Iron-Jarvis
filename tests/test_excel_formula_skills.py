"""v1.90.0: formula intelligence + disposable code + agent-authored skills.

* The formula engine evaluates the financial-statement subset against real
  sheet values — validation is COMPUTED, never assumed.
* excel_formula_check proves a formula against the sheet it came from.
* excel_sheet_spec / excel_apply_spec reproduce a sheet's formulas AND
  formatting, then validate the reproduction.
* excel_accounts_diff notices added/removed/moved accounts so label-anchored
  formulas get updated (and re-proven) instead of silently drifting.
* run_code executes disposable scripts (kept only on request).
* skill_create persists a proven approach as a live, searchable skill.

Offline throughout.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from iron_jarvis.daemon.app import create_app
from iron_jarvis.documents.excel_formula import (
    FormulaError,
    Grid,
    evaluate_formula,
)
from iron_jarvis.documents.tools import document_tools
from iron_jarvis.tools.base import ToolContext


def _ctx(ws: Path) -> ToolContext:
    return ToolContext(
        workspace=ws, session_id="t", agent_run_id="t",
        config=None, event_bus=None, engine=None,
    )


def _tool(name: str):
    return next(t for t in document_tools() if t.name == name)


def _statement(path: Path) -> None:
    """A mini P&L: labels in A, amounts in B, a literal total in B6."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PL"
    rows = [("Account", "Amount"), ("Revenue", 5000), ("COGS", -2000),
            ("Payroll", -1500), ("Rent", -500)]
    for r in rows:
        ws.append(r)
    ws["A6"] = "Net income"
    ws["B6"] = 1000  # literal — the ground truth a derived formula must hit
    ws["B7"] = "=SUM(B2:B5)"  # a stored formula (no cached value via openpyxl)
    ws["B2"].number_format = "#,##0.00"
    from openpyxl.styles import Font

    ws["A1"].font = Font(bold=True)
    wb.create_sheet("Notes")["A1"] = "memo"
    wb.save(str(path))


# --- the formula engine -------------------------------------------------------


def test_evaluator_core(tmp_path):
    _statement(tmp_path / "pl.xlsx")
    grid = Grid(tmp_path / "pl.xlsx")
    try:
        assert evaluate_formula("=SUM(B2:B5)", grid, "PL") == 1000
        assert evaluate_formula("=B2+B3", grid, "PL") == 3000
        assert evaluate_formula("=ROUND(B2/3, 2)", grid, "PL") == 1666.67
        assert evaluate_formula("=IF(SUM(B2:B5)>0, 1, 0)", grid, "PL") == 1
        assert evaluate_formula("=ABS(B3)", grid, "PL") == 2000
        assert evaluate_formula("=AVERAGE(B2:B5)", grid, "PL") == 250
        assert evaluate_formula("=SUBTOTAL(9, B2:B5)", grid, "PL") == 1000
        assert evaluate_formula("=10%", grid, "PL") == pytest.approx(0.1)
        assert evaluate_formula("=2^3*2", grid, "PL") == 16
        # Sheet-qualified refs + blanks-as-zero + text skipped by SUM.
        assert evaluate_formula("=PL!B2", grid, "Notes") == 5000
        assert evaluate_formula("=Z99+5", grid, "PL") == 5  # blank cell = 0
        assert evaluate_formula("=SUM(A1:B5)", grid, "PL") == 1000  # labels skipped
        with pytest.raises(FormulaError, match="VLOOKUP"):
            evaluate_formula("=VLOOKUP(A1,B1:C5,2)", grid, "PL")
        with pytest.raises(FormulaError, match="named ranges"):
            evaluate_formula("=SUM(B:B)", grid, "PL")  # whole-column unsupported
        with pytest.raises(FormulaError, match="division by zero"):
            evaluate_formula("=1/0", grid, "PL")
    finally:
        grid.close()


# --- excel_formula_check ------------------------------------------------------


async def test_formula_check_validates_against_the_sheet(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    _statement(ws / "pl.xlsx")
    check = _tool("excel_formula_check")
    res = await check.execute(
        {"path": "pl.xlsx", "sheet": "PL", "formula": "=SUM(B2:B5)",
         "expect_cell": "B6"}, _ctx(ws)
    )
    assert res.ok, res.error
    assert res.data["match"] is True and res.data["computed"] == 1000
    assert "MATCH" in res.output
    # A WRONG formula (range misses a row) is caught with both numbers shown.
    res = await check.execute(
        {"path": "pl.xlsx", "sheet": "PL", "formula": "=SUM(B2:B4)",
         "expect_cell": "B6"}, _ctx(ws)
    )
    assert res.ok and res.data["match"] is False
    assert "MISMATCH" in res.output
    # cell= mode reads the STORED formula; openpyxl-authored files carry no
    # cached value, and the tool says so instead of inventing a comparison.
    res = await check.execute(
        {"path": "pl.xlsx", "sheet": "PL", "cell": "B7"}, _ctx(ws)
    )
    assert res.ok and res.data["computed"] == 1000
    assert "no cached value" in res.output
    # Unsupported function → honest refusal, not a wrong number.
    res = await check.execute(
        {"path": "pl.xlsx", "sheet": "PL", "formula": "=XLOOKUP(1,A:A,B:B)"},
        _ctx(ws),
    )
    assert not res.ok and "not supported" in res.error


# --- sheet spec: capture + reproduce + validate -------------------------------


async def test_spec_roundtrip_reproduces_formulas_and_formatting(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    _statement(ws / "pl.xlsx")
    spec_res = await _tool("excel_sheet_spec").execute(
        {"path": "pl.xlsx", "sheet": "PL"}, _ctx(ws)
    )
    assert spec_res.ok, spec_res.error
    spec = spec_res.data
    assert spec["cells"]["B7"]["formula"] == "=SUM(B2:B5)"
    assert spec["cells"]["B2"]["number_format"] == "#,##0.00"
    assert spec["cells"]["A1"]["bold"] is True
    assert spec["labels"]["2"] == "Revenue" and spec["labels"]["6"] == "Net income"

    # Reproduce onto a FRESH workbook that carries the same data rows.
    tgt = ws / "repro.xlsx"
    wb = Workbook()
    s = wb.active
    s.title = "PL"
    for row in [("Account", "Amount"), ("Revenue", 5000), ("COGS", -2000),
                ("Payroll", -1500), ("Rent", -500), ("Net income", 1000)]:
        s.append(row)
    wb.save(str(tgt))
    apply_res = await _tool("excel_apply_spec").execute(
        {"path": "repro.xlsx", "spec": spec}, _ctx(ws)
    )
    assert apply_res.ok, apply_res.error
    assert apply_res.data["applied"]["formulas"] == 1
    # The validation COMPUTED the written formula against the target's data.
    check = next(c for c in apply_res.data["validation"] if c["cell"] == "B7")
    assert check["computed"] == 1000
    out = load_workbook(str(tgt))
    pl = out["PL"]
    assert pl["B7"].value == "=SUM(B2:B5)"
    assert pl["B2"].number_format == "#,##0.00"
    assert bool(pl["A1"].font.bold) is True


async def test_apply_spec_verify_against_source_flags_drift(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    spec = {
        "sheet": "PL",
        "cells": {"B3": {"formula": "=SUM(B1:B2)", "value": 999.0}},  # wrong on purpose
    }
    tgt = ws / "t.xlsx"
    wb = Workbook()
    s = wb.active
    s.title = "PL"
    s["B1"], s["B2"] = 10, 20
    wb.save(str(tgt))
    res = await _tool("excel_apply_spec").execute(
        {"path": "t.xlsx", "spec": spec, "verify_against_source": True}, _ctx(ws)
    )
    assert res.ok
    assert res.data["failed"] == 1  # computed 30 vs recorded 999 — flagged
    assert "VALIDATION FAILED" in res.output


# --- accounts diff ------------------------------------------------------------


async def test_accounts_diff_sees_added_removed_moved(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    _statement(ws / "old.xlsx")
    wb = Workbook()
    s = wb.active
    s.title = "PL"
    # Marketing inserted; Rent removed; Net income shifted up a row.
    for row in [("Account",), ("Revenue",), ("COGS",), ("Marketing",),
                ("Payroll",), ("Net income",)]:
        s.append(row)
    wb.save(str(ws / "new.xlsx"))
    res = await _tool("excel_accounts_diff").execute(
        {"path_a": "old.xlsx", "sheet_a": "PL", "path_b": "new.xlsx"}, _ctx(ws)
    )
    assert res.ok, res.error
    assert res.data["added"] == [{"label": "Marketing", "row": 4}]
    assert {r["label"] for r in res.data["removed"]} == {"Rent"}
    assert any(m["label"] == "Payroll" and m["from"] == 4 and m["to"] == 5
               for m in res.data["moved"])
    assert "excel_formula_check" in res.output  # the tool teaches the workflow


# --- run_code: disposable scripts ---------------------------------------------


async def test_run_code_python_disposable_and_kept(tmp_path):
    from iron_jarvis.tools.runcode import RunCodeTool

    ws = tmp_path / "ws"
    ws.mkdir()
    tool = RunCodeTool()
    res = await tool.execute(
        {"language": "python", "code": "print(sum(range(101)))"}, _ctx(ws)
    )
    assert res.ok, res.error
    assert "5050" in res.output and "script discarded" in res.output
    assert not (ws / ".scratch").exists() or not any((ws / ".scratch").iterdir())
    res = await tool.execute(
        {"language": "python", "code": "print('kept')", "keep": True,
         "filename": "solver"}, _ctx(ws)
    )
    assert res.ok and res.data["kept"] == "scripts/solver.py"
    assert (ws / "scripts" / "solver.py").is_file()
    # A failing script is an HONEST failure, not a fabricated success.
    res = await tool.execute(
        {"language": "python", "code": "raise SystemExit(3)"}, _ctx(ws)
    )
    assert not res.ok and "exited 3" in (res.error or "")


# --- skill_create: keep the proven solution -----------------------------------


async def test_skill_create_persists_and_is_live(tmp_path):
    client = TestClient(create_app(str(tmp_path)))
    platform = client.app.state.platform
    tool = platform.registry.get("skill_create")
    assert tool is not None
    res = await tool.execute(
        {
            "name": "pl-net-income-check",
            "description": "Validate a P&L net-income formula against the sheet",
            "instructions": (
                "1. excel_profile the statement.\n"
                "2. Derive =SUM over the amount column between the first "
                "account row and the last.\n"
                "3. excel_formula_check it against the stated total.\n"
                "4. Only then excel_edit it in."
            ),
        },
        _ctx(tmp_path),
    )
    assert res.ok, res.error
    sk = platform.skills.get("pl-net-income-check")
    assert sk is not None and "excel_formula_check" in sk.instructions
    names = [s["name"] for s in client.get("/skills").json()["skills"]]
    assert "pl-net-income-check" in names


# --- registration + permissions ----------------------------------------------


def test_new_tools_registered_with_sane_permissions(tmp_path):
    client = TestClient(create_app(str(tmp_path)))
    platform = client.app.state.platform
    names = set(platform.registry.names())
    for n in ("excel_formula_check", "excel_sheet_spec", "excel_apply_spec",
              "excel_accounts_diff", "run_code", "skill_create"):
        assert n in names, n
    perms = platform.config.permissions
    assert perms["excel_formula_check"] == "allow"
    assert perms["run_code"] == "ask"  # code execution stays consent-gated
    assert perms["skill_create"] == "ask"  # skills inject into future prompts
