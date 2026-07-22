"""Excel parity tools (v1.78.0): structured reads + in-place edits.

The capability that made "works with Excel" magic, model-agnostic: read a
workbook as structure (sheets/ranges/formulas), edit an existing workbook in
place (values, formulas, new sheets) with untouched formatting preserved and
TX-01 undo capturing the exact prior bytes.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from iron_jarvis.documents.tools import document_tools
from iron_jarvis.tools.base import ToolContext


def _ctx(ws: Path) -> ToolContext:
    return ToolContext(
        workspace=ws, session_id="t", agent_run_id="t",
        config=None, event_bus=None, engine=None,
    )


def _tool(name: str):
    return next(t for t in document_tools() if t.name == name)


def _book(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"
    ws["A1"] = "Client"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "Billed"
    ws["A2"] = "Acme"
    ws["B2"] = 1200
    ws["A3"] = "Birch"
    ws["B3"] = 800
    wb.create_sheet("Notes")
    wb.save(str(path))


async def test_read_overview_then_range(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    _book(ws / "clients.xlsx")
    tool = _tool("excel_read")
    res = await tool.execute({"path": "clients.xlsx"}, _ctx(ws))
    assert res.ok, res.error
    assert res.data["sheets"] == ["Q3", "Notes"]
    assert res.data["overview"][0] == {"sheet": "Q3", "rows": 3, "cols": 2}

    res = await tool.execute(
        {"path": "clients.xlsx", "sheet": "Q3", "range": "A1:B3"}, _ctx(ws)
    )
    assert res.ok
    assert res.data["rows"] == [["Client", "Billed"], ["Acme", 1200], ["Birch", 800]]


async def test_edit_in_place_preserves_untouched_formatting(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    path = ws / "clients.xlsx"
    _book(path)
    tool = _tool("excel_edit")
    res = await tool.execute(
        {
            "path": "clients.xlsx",
            "sheet": "Q3",
            "edits": [
                {"cell": "B2", "value": 1500},
                {"cell": "B4", "formula": "=SUM(B2:B3)"},
                {"cell": "A1", "value": "Client name", "sheet": "Q3"},
            ],
            "add_sheets": ["Summary"],
        },
        _ctx(ws),
    )
    assert res.ok, res.error
    assert res.data["applied"] == 3
    wb = load_workbook(str(path))
    q3 = wb["Q3"]
    assert q3["B2"].value == 1500
    assert q3["B4"].value == "=SUM(B2:B3)"
    assert q3["A1"].value == "Client name"
    assert q3["A1"].font.bold is True  # untouched formatting survived the edit
    assert "Summary" in wb.sheetnames


async def test_edit_undo_restores_exact_prior_bytes(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    path = ws / "clients.xlsx"
    _book(path)
    before = path.read_bytes()
    tool = _tool("excel_edit")

    class _Cfg:  # make_file_descriptor stores the pre-image under config.home
        home = ws

    ctx = ToolContext(
        workspace=ws, session_id="t", agent_run_id="t",
        config=_Cfg(), event_bus=None, engine=None,
    )
    undo = await tool.capture_undo(
        {"path": "clients.xlsx", "edits": [{"cell": "B2", "value": 9}]}, ctx
    )
    assert undo is not None and undo["kind"] == "file_restore"
    res = await tool.execute(
        {"path": "clients.xlsx", "edits": [{"cell": "B2", "value": 9}]}, ctx
    )
    assert res.ok
    assert path.read_bytes() != before
    rev = await tool.revert(undo, ctx)
    assert rev.ok, rev.error
    assert path.read_bytes() == before  # byte-exact restore


async def test_edit_is_workspace_confined_and_honest(tmp_path):
    outside = tmp_path / "outside.xlsx"
    _book(outside)
    ws = tmp_path / "ws"
    ws.mkdir()
    tool = _tool("excel_edit")
    res = await tool.execute(
        {"path": str(outside), "edits": [{"cell": "A1", "value": "x"}]}, _ctx(ws)
    )
    assert not res.ok  # absolute escape refused by safe_path
    res = await tool.execute({"path": "missing.xlsx", "edits": [{"cell": "A1", "value": 1}]}, _ctx(ws))
    assert not res.ok and "write_document" in (res.error or "")
    res = await tool.execute({"path": "clients.xlsx"}, _ctx(ws))
    assert not res.ok  # nothing to do is an error, not a silent no-op


async def test_read_formulas_mode(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    path = ws / "f.xlsx"
    wb = Workbook()
    wb.active["A1"] = 2
    wb.active["A2"] = "=A1*10"
    wb.save(str(path))
    tool = _tool("excel_read")
    res = await tool.execute(
        {"path": "f.xlsx", "range": "A1:A2", "include_formulas": True}, _ctx(ws)
    )
    assert res.ok
    assert res.data["rows"] == [[2], ["=A1*10"]]
