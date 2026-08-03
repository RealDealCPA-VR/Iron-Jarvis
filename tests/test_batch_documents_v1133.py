"""Batch document pipeline (v1.133.0, Wave 3).

Folder → per-document extraction → synthesis over the EXTRACTIONS. These tests
pin the load-bearing guarantees, all offline via a scripted fake router:

* sweep honors the shared fs policy and records every exclusion (denied,
  unsupported, over-max_files) — nothing silently dropped;
* per-doc extraction JSON is persisted with the content hash and RESUMES
  (cached, LLM not called again) for unchanged files;
* a malformed reply gets exactly one repair round; a double failure lands in
  failed[] and the batch continues;
* the synthesis input is built from EXTRACTIONS, never raw document text;
* the docx/xlsx deliverables are created by the REAL writers, and excluded
  documents are listed in them by code, not by model compliance.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

from iron_jarvis.core import fs_policy
from iron_jarvis.documents import batch
from iron_jarvis.documents.batch import extract_one, run_batch, sweep
from iron_jarvis.documents.readers import extract_text
from iron_jarvis.documents.tools import document_tools
from iron_jarvis.providers.adapters.base import LLMResponse
from iron_jarvis.providers.router import RouteResult
from iron_jarvis.tools.base import ToolContext

RAW_MARKER = "RAWMARKER-XYZZY-9001"  # seeded in a doc body, dropped by extraction


class ScriptedRouter:
    """Replays scripted completion texts in order; records every prompt.

    An ``Exception`` item raises instead (a real-provider failure)."""

    def __init__(self, replies, provider="anthropic"):
        self.replies = list(replies)
        self.provider = provider
        self.calls: list[tuple[str, str]] = []  # (system, user content)

    async def complete(self, *, system, messages, tools, task_class=None, **kw):
        assert len(messages) == 1 and messages[0].role == "user"
        self.calls.append((system, messages[0].content))
        assert self.replies, "router called more times than the script allows"
        item = self.replies.pop(0)
        if isinstance(item, Exception):
            raise item
        return RouteResult(LLMResponse(text=item), self.provider, "test-model")


def _ext_reply(summary, facts=()):
    return json.dumps(
        {
            "summary": summary,
            "facts": list(facts),
            "entities": {"people": [], "orgs": [], "dates": [], "amounts": []},
            "figures": [],
        }
    )


MD_REPLY = "# Batch Report\n\n- combined finding one"
SHEETS_REPLY = json.dumps(
    {"sheets": {"Overview": [["Document", "Summary"], ["a.txt", "Alpha doc"]]}}
)


def _docs(folder: Path, **bodies: str) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    for name, body in bodies.items():
        (folder / name.replace("_", ".")).write_text(body, encoding="utf-8")


# ------------------------------------------------------------------ sweep ----


def test_sweep_records_denied_and_unsupported(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha", denied_txt="secret")
    (src / "blob.zip").write_bytes(b"\x00\x01")
    denied = src / "denied.txt"
    fs_policy.register_protected_root(denied)
    try:
        files, skipped = sweep(src, max_files=25)
    finally:  # never leak a protected root into other tests
        fs_policy._PROTECTED_ROOTS.discard(fs_policy._canonical(denied))
    assert [p.name for p in files] == ["a.txt"]
    reasons = {Path(s["file"]).name: s["reason"] for s in skipped}
    assert "read denied" in reasons["denied.txt"]
    assert "unsupported" in reasons["blob.zip"]


def test_sweep_max_files_truncation_recorded(tmp_path):
    src = tmp_path / "docs"
    _docs(src, **{f"doc{i}_txt": f"body {i}" for i in range(5)})
    files, skipped = sweep(src, max_files=3)
    # deterministic name order, truncated at the limit
    assert [p.name for p in files] == ["doc0.txt", "doc1.txt", "doc2.txt"]
    over = [s for s in skipped if "max_files" in s["reason"]]
    assert {Path(s["file"]).name for s in over} == {"doc3.txt", "doc4.txt"}


def test_sweep_records_subfolders_not_descended(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha")
    (src / "nested").mkdir()
    (src / "nested" / "inner.txt").write_text("hidden", encoding="utf-8")
    files, skipped = sweep(src, max_files=25)
    assert [p.name for p in files] == ["a.txt"]  # never descended
    sub = [s for s in skipped if Path(s["file"]).name == "nested"]
    assert len(sub) == 1 and "subfolder" in sub[0]["reason"]


def test_slug_distinct_for_same_name_in_different_folders(tmp_path):
    a = tmp_path / "one" / "report.txt"
    b = tmp_path / "two" / "report.txt"
    assert batch.slug_for(a) != batch.slug_for(b)  # path digest disambiguates
    assert batch.slug_for(a) == batch.slug_for(a)  # and is stable across runs


# ----------------------------------------------------- persist + resume ------


async def test_extraction_persisted_with_hash_and_resume(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body", b_txt="bravo body")
    out = tmp_path / "out"

    r1 = ScriptedRouter([_ext_reply("Alpha"), _ext_reply("Bravo"), MD_REPLY])
    res1 = await run_batch(src, out, r1, output="docx")
    assert (res1["processed"], res1["cached"], res1["failed"]) == (2, 0, [])
    records = sorted((out / "extractions").glob("*.json"))
    assert len(records) == 2
    rec = json.loads(records[0].read_text(encoding="utf-8"))
    assert rec["source"].endswith("a.txt")
    assert rec["sha256"] == hashlib.sha256(b"alpha body").hexdigest()
    assert rec["size"] == len(b"alpha body") and rec["mtime"] > 0
    assert rec["extraction"]["summary"] == "Alpha"

    # RESUME: unchanged files are cached — the LLM sees only the synthesis call.
    r2 = ScriptedRouter([MD_REPLY])
    res2 = await run_batch(src, out, r2, output="docx")
    assert (res2["processed"], res2["cached"]) == (0, 2)
    assert len(r2.calls) == 1  # the ONLY call is synthesis, not re-extraction
    assert "Per-document extractions" in r2.calls[0][1]

    # A CHANGED file re-extracts; the untouched one stays cached.
    (src / "b.txt").write_text("bravo body v2", encoding="utf-8")
    r3 = ScriptedRouter([_ext_reply("Bravo v2"), MD_REPLY])
    res3 = await run_batch(src, out, r3, output="docx")
    assert (res3["processed"], res3["cached"]) == (1, 2 - 1)


async def test_corrupt_record_reextracts_and_mtime_is_untrusted(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    out = tmp_path / "out"
    r1 = ScriptedRouter([_ext_reply("Alpha"), MD_REPLY])
    await run_batch(src, out, r1, output="docx")
    record = next((out / "extractions").glob("*.json"))

    # A corrupt record must re-extract, never crash the batch.
    record.write_text("{not json", encoding="utf-8")
    r2 = ScriptedRouter([_ext_reply("Alpha again"), MD_REPLY])
    res = await run_batch(src, out, r2, output="docx")
    assert (res["processed"], res["cached"], res["failed"]) == (1, 0, [])

    # Same name, same size, SAME mtime, different bytes → hash mismatch wins.
    import os

    st = (src / "a.txt").stat()
    (src / "a.txt").write_text("delta body", encoding="utf-8")  # same length
    os.utime(src / "a.txt", (st.st_atime, st.st_mtime))
    r3 = ScriptedRouter([_ext_reply("Delta"), MD_REPLY])
    res = await run_batch(src, out, r3, output="docx")
    assert (res["processed"], res["cached"]) == (1, 0)


# ------------------------------------------------------ repair + failure -----


async def test_malformed_extraction_gets_one_repair(tmp_path):
    doc = tmp_path / "a.txt"
    doc.write_text("alpha body", encoding="utf-8")
    router = ScriptedRouter(["utterly not json", _ext_reply("Alpha", ["f1"])])
    ex = await extract_one(doc, router)
    assert ex["summary"] == "Alpha" and ex["facts"] == ["f1"]
    assert len(router.calls) == 2
    repair_user = router.calls[1][1]
    # the repair prompt feeds the validation error + the rejected reply back
    assert "rejected" in repair_user and "JSON" in repair_user
    assert "utterly not json" in repair_user


async def test_double_failure_lands_in_failed_and_batch_continues(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body", b_txt="bravo body")
    out = tmp_path / "out"
    router = ScriptedRouter(
        ["junk one", "junk two", _ext_reply("Bravo", ["b fact"]), MD_REPLY, SHEETS_REPLY]
    )
    res = await run_batch(src, out, router, output="both")
    assert res["processed"] == 1
    assert len(res["failed"]) == 1 and res["failed"][0]["file"].endswith("a.txt")
    assert res["failed"][0]["error"]
    # the batch continued: BOTH deliverables exist, honestly listing a.txt
    assert len(res["deliverables"]) == 2 and res["synthesis_errors"] == []
    docx_text = extract_text(out / "synthesis.docx")
    assert "Documents not included" in docx_text and "a.txt" in docx_text
    from openpyxl import load_workbook

    wb = load_workbook(out / "synthesis.xlsx")
    try:
        assert "Not included" in wb.sheetnames
        rows = list(wb["Not included"].iter_rows(values_only=True))
        assert rows[0] == ("File", "Reason")
        assert rows[1][0] == "a.txt" and "extraction failed" in rows[1][1]
    finally:
        wb.close()


async def test_provider_failure_is_the_docs_error_never_fabricated(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    out = tmp_path / "out"
    router = ScriptedRouter([RuntimeError("provider down")])
    res = await run_batch(src, out, router, output="docx")
    assert res["processed"] == 0 and res["deliverables"] == []
    assert "provider down" in res["failed"][0]["error"]
    # with zero extractions there is nothing honest to synthesize from
    assert "nothing to synthesize" in res["synthesis_errors"][0]["error"]
    assert not list((out / "extractions").glob("*.json"))


async def test_mock_provider_never_extracts(tmp_path):
    doc = tmp_path / "a.txt"
    doc.write_text("alpha body", encoding="utf-8")
    router = ScriptedRouter([_ext_reply("Fabricated")], provider="mock")
    try:
        await extract_one(doc, router)
        raise AssertionError("mock extraction must raise")
    except RuntimeError as exc:
        assert "mock" in str(exc)


# --------------------------------------------------- synthesis boundaries ----


async def test_synthesis_input_is_extractions_not_raw_text(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt=f"alpha body containing {RAW_MARKER} secret detail")
    out = tmp_path / "out"
    router = ScriptedRouter(
        [_ext_reply("Alpha report about revenue", ["Revenue was 100"]), MD_REPLY]
    )
    res = await run_batch(src, out, router, output="docx")
    assert res["processed"] == 1
    extract_call, synth_call = router.calls
    assert RAW_MARKER in extract_call[1]  # extraction DOES see the raw doc
    # synthesis sees the extraction's content and never the raw text
    assert "Revenue was 100" in synth_call[1]
    assert "Alpha report about revenue" in synth_call[1]
    assert RAW_MARKER not in synth_call[0] and RAW_MARKER not in synth_call[1]


async def test_deliverables_created_by_real_writers_and_shape(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    out = tmp_path / "out"
    router = ScriptedRouter([_ext_reply("Alpha"), MD_REPLY, SHEETS_REPLY])
    res = await run_batch(src, out, router, output="both")
    assert set(res) >= {
        "processed", "cached", "failed", "skipped", "deliverables",
        "synthesis_errors", "extraction_dir",
    }
    assert (res["processed"], res["cached"], res["failed"], res["skipped"]) == (
        1, 0, [], [],
    )
    docx, xlsx = (Path(p) for p in res["deliverables"])
    assert docx.is_file() and docx.suffix == ".docx"
    assert xlsx.is_file() and xlsx.suffix == ".xlsx"
    assert "Batch Report" in extract_text(docx)  # a real Word document
    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    try:
        ws = wb["Overview"]
        assert ws["A1"].value == "Document" and ws["A2"].value == "a.txt"
    finally:
        wb.close()


async def test_bad_sheet_spec_gets_one_repair(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    out = tmp_path / "out"
    router = ScriptedRouter([_ext_reply("Alpha"), "not a sheets object", SHEETS_REPLY])
    res = await run_batch(src, out, router, output="xlsx")
    assert res["deliverables"] and res["synthesis_errors"] == []
    assert len(router.calls) == 3  # extract + bad synth + repaired synth


async def test_mock_provider_never_synthesizes(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    out = tmp_path / "out"
    # Run 1 with a real provider extracts + delivers.
    await run_batch(src, out, ScriptedRouter([_ext_reply("Alpha"), MD_REPLY]), output="docx")
    # Run 2 resumes from cache with ONLY the mock connected: the synthesis
    # call site must refuse too — never a mock-fabricated deliverable.
    res = await run_batch(src, out, ScriptedRouter([MD_REPLY], provider="mock"), output="docx")
    assert (res["processed"], res["cached"]) == (0, 1)
    assert res["deliverables"] == []
    assert len(res["synthesis_errors"]) == 1
    assert "mock" in res["synthesis_errors"][0]["error"]


def test_validate_sheets_rejects_silent_data_loss_shapes():
    # NaN/Infinity pass json.loads and openpyxl writes them as EMPTY cells;
    # >32,767-char cells get silently TRUNCATED by openpyxl; control chars
    # crash the writer after validation. All three must fail validation so
    # the repair round fires instead.
    for label, raw in [
        ("nan", '{"sheets": {"S": [["h"], [NaN]]}}'),
        ("inf", '{"sheets": {"S": [["h"], [Infinity]]}}'),
        ("huge", json.dumps({"sheets": {"S": [["h"], ["x" * 40_000]]}})),
        ("ctrl", json.dumps({"sheets": {"S": [["h"], ["bad\x01char"]]}})),
    ]:
        try:
            batch._validate_sheets(batch._parse_json_object(raw))
            raise AssertionError(f"{label} must be rejected")
        except ValueError as exc:
            assert "sheet" in str(exc), (label, exc)
    # ...while honest shapes the writer supports still pass: None cells,
    # ragged rows, bools, formula strings.
    spec = batch._validate_sheets(
        {"sheets": {"S": [["a", "b"], [None], [True, "=SUM(A1:A2)", 3.5]]}}
    )
    assert spec["sheets"]["S"][2][1] == "=SUM(A1:A2)"


async def test_nan_sheet_reply_gets_repair_then_real_workbook(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    out = tmp_path / "out"
    nan_reply = '{"sheets": {"Overview": [["Doc", "Total"], ["a.txt", NaN]]}}'
    router = ScriptedRouter([_ext_reply("Alpha"), nan_reply, SHEETS_REPLY])
    res = await run_batch(src, out, router, output="xlsx")
    assert res["deliverables"] and res["synthesis_errors"] == []
    assert len(router.calls) == 3  # extract + rejected NaN synth + repair
    assert "non-finite" in router.calls[2][1]  # repair names the defect


async def test_oversized_doc_is_clipped_for_extraction(tmp_path):
    doc = tmp_path / "big.txt"
    doc.write_text("word " * 10_000, encoding="utf-8")  # 50k chars
    router = ScriptedRouter([_ext_reply("Big")])
    await extract_one(doc, router)
    user = router.calls[0][1]
    assert len(user) < batch.MAX_DOC_CHARS + 500  # bounded input, always
    assert "truncated" in user  # and the clip is disclosed to the model


# ------------------------------------------------------------- the tool ------


def _ctx(ws: Path) -> ToolContext:
    ws.mkdir(parents=True, exist_ok=True)
    return ToolContext(
        workspace=ws, session_id="t", agent_run_id="t",
        config=None, event_bus=None, engine=None,
    )


async def test_batch_documents_tool_end_to_end(tmp_path):
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha body")
    router = ScriptedRouter([_ext_reply("Alpha"), MD_REPLY])
    tool = next(
        t for t in document_tools(router_resolver=lambda: router)
        if t.name == "batch_documents"
    )
    ctx = _ctx(tmp_path / "ws")
    res = await tool.execute(
        {"folder": str(src), "output": "docx", "instructions": "summarize"}, ctx
    )
    assert res.ok, res.error
    assert res.data["processed"] == 1
    # deliverables + extraction cache live INSIDE the session workspace
    rel = res.data["deliverables"][0]
    assert not Path(rel).is_absolute()
    assert (ctx.workspace / rel).is_file()
    assert Path(res.data["extraction_dir"]).resolve().is_relative_to(
        ctx.workspace.resolve()
    )
    assert "1 extracted" in res.output


async def test_batch_documents_tool_rejects_bad_args(tmp_path):
    router = ScriptedRouter([])
    tool = next(
        t for t in document_tools(router_resolver=lambda: router)
        if t.name == "batch_documents"
    )
    ctx = _ctx(tmp_path / "ws")
    res = await tool.execute({"folder": str(tmp_path / "missing")}, ctx)
    assert not res.ok and "not a folder" in res.error
    src = tmp_path / "docs"
    _docs(src, a_txt="alpha")
    res = await tool.execute({"folder": str(src), "output": "pptx"}, ctx)
    assert not res.ok and "unknown output" in res.error
    # a bare factory (no router wired) reports honestly instead of crashing
    bare = next(t for t in document_tools() if t.name == "batch_documents")
    res = await bare.execute({"folder": str(src)}, ctx)
    assert not res.ok and "router" in res.error


async def test_tool_denies_protected_folder_as_error_not_exception(tmp_path):
    src = tmp_path / "vault"
    _docs(src, a_txt="secret")
    fs_policy.register_protected_root(src)
    try:
        tool = next(
            t for t in document_tools(router_resolver=lambda: ScriptedRouter([]))
            if t.name == "batch_documents"
        )
        res = await tool.execute({"folder": str(src)}, _ctx(tmp_path / "ws"))
    finally:
        fs_policy._PROTECTED_ROOTS.discard(fs_policy._canonical(src))
    assert not res.ok and "read denied" in res.error  # recorded, never raised


async def test_tool_integration_mixed_types_both_formats(tmp_path):
    """The real factory, txt+md+csv in one folder, both deliverables read back."""
    src = tmp_path / "docs"
    src.mkdir()
    (src / "alpha.txt").write_text("alpha revenue was 100", encoding="utf-8")
    (src / "bravo.md").write_text("# Bravo\n\ncosts were 40", encoding="utf-8")
    (src / "charlie.csv").write_text("item,amount\nfee,60\n", encoding="utf-8")
    router = ScriptedRouter(
        [
            _ext_reply("Alpha doc", ["revenue 100"]),
            _ext_reply("Bravo doc", ["costs 40"]),
            _ext_reply("Charlie doc", ["fee 60"]),
            MD_REPLY,
            SHEETS_REPLY,
        ]
    )
    tools = document_tools(router_resolver=lambda: router)
    tool = next(t for t in tools if t.name == "batch_documents")
    # the registry-visible contract agents will see
    spec = tool.spec()
    assert spec["name"] == "batch_documents"
    assert set(spec["input_schema"]["properties"]) == {
        "folder", "instructions", "output", "max_files",
    }
    assert spec["input_schema"]["required"] == ["folder"]
    for claim in ("top level only", "12k", "RESUMES"):
        assert claim in spec["description"], claim
    assert tool.returns_untrusted_content is True

    ctx = _ctx(tmp_path / "ws")
    res = await tool.execute({"folder": str(src), "output": "both"}, ctx)
    assert res.ok, res.error
    assert res.data["processed"] == 3 and res.data["failed"] == []
    assert len(res.data["deliverables"]) == 2
    for rel in res.data["deliverables"]:
        assert "\\" not in rel  # workspace-relative, forward slashes
        assert (ctx.workspace / rel).is_file()
    docx = next(p for p in res.data["deliverables"] if p.endswith(".docx"))
    assert "Batch Report" in extract_text(ctx.workspace / docx)
    from openpyxl import load_workbook

    xlsx = next(p for p in res.data["deliverables"] if p.endswith(".xlsx"))
    wb = load_workbook(ctx.workspace / xlsx)
    try:
        assert wb["Overview"]["A1"].value == "Document"
    finally:
        wb.close()
