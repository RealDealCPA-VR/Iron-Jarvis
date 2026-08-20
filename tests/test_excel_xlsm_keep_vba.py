"""A macro-enabled workbook edited IN PLACE keeps its macros.

`excel_edit` and `excel_apply_spec` both admit `.xlsm` and both `wb.save()`
back over the SAME path. openpyxl drops the `xl/vbaProject.bin` part unless
`keep_vba=True`, so an in-place save without it silently destroys every macro
in a client workbook — and because the package no longer matches the
macro-enabled extension, Excel commonly refuses to open the result at all —
while the tool reports a clean success. `keep_vba` must track the suffix at
BOTH save-in-place sites.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook

from iron_jarvis.documents.tools import document_tools
from iron_jarvis.tools.base import ToolContext

VBA_PART = "xl/vbaProject.bin"
VBA_BYTES = b"FAKE-VBA-PROJECT"


def _ctx(ws: Path) -> ToolContext:
    return ToolContext(
        workspace=ws, session_id="t", agent_run_id="t",
        config=None, event_bus=None, engine=None,
    )


def _tool(name: str):
    return next(t for t in document_tools() if t.name == name)


def _macro_book(path: Path) -> None:
    """Write a real .xlsm: an openpyxl workbook plus a vbaProject part.

    openpyxl cannot author VBA, so the part is injected into the zip the way a
    macro-enabled workbook carries it (content-type override + the binary).
    That is exactly what `keep_vba` preserves or discards.
    """
    plain = path.with_suffix(".seed.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"
    ws["A1"] = "Client"
    ws["B1"] = 1200
    wb.save(str(plain))

    with zipfile.ZipFile(plain) as src, zipfile.ZipFile(path, "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/vbaProject.bin" '
                    b'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
                )
            dst.writestr(item.filename, data)
        dst.writestr(VBA_PART, VBA_BYTES)
    plain.unlink()
    assert VBA_PART in _parts(path)


def _parts(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        return z.namelist()


async def test_excel_edit_in_place_keeps_the_vba_project(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    book = ws / "budget.xlsm"
    _macro_book(book)

    res = await _tool("excel_edit").execute(
        {"path": "budget.xlsm", "sheet": "Q3",
         "edits": [{"cell": "B2", "value": 99}]},
        _ctx(ws),
    )
    assert res.ok, res.error
    assert res.data["applied"] == 1
    # The edit landed AND the macros survived the in-place save.
    assert VBA_PART in _parts(book)
    with zipfile.ZipFile(book) as z:
        assert z.read(VBA_PART) == VBA_BYTES


async def test_excel_apply_spec_on_existing_xlsm_keeps_the_vba_project(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    book = ws / "budget.xlsm"
    _macro_book(book)

    res = await _tool("excel_apply_spec").execute(
        {"path": "budget.xlsm", "sheet": "Q3",
         "spec": {"sheet": "Q3", "cells": {"B3": {"formula": "=B1*2"}}}},
        _ctx(ws),
    )
    assert res.ok, res.error
    assert VBA_PART in _parts(book)
    with zipfile.ZipFile(book) as z:
        assert z.read(VBA_PART) == VBA_BYTES


async def test_keep_vba_tracks_the_suffix_at_both_in_place_sites(tmp_path, monkeypatch):
    """The flag is derived from the extension, not hardcoded either way."""
    import openpyxl

    real = openpyxl.load_workbook
    seen: list[object] = []

    def spy(*args, **kwargs):
        if "keep_vba" in kwargs:
            seen.append(kwargs["keep_vba"])
        return real(*args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", spy)

    ws = tmp_path / "ws"
    ws.mkdir()
    plain = ws / "plain.xlsx"
    wb = Workbook()
    wb.active.title = "Q3"
    wb.active["A1"] = 1
    wb.save(str(plain))
    macro = ws / "macro.xlsm"
    _macro_book(macro)

    edit = _tool("excel_edit")
    spec_tool = _tool("excel_apply_spec")
    spec = {"sheet": "Q3", "cells": {"C1": {"formula": "=A1"}}}

    for name, expected in (("plain.xlsx", False), ("macro.xlsm", True)):
        seen.clear()
        res = await edit.execute(
            {"path": name, "sheet": "Q3", "edits": [{"cell": "B2", "value": 1}]},
            _ctx(ws),
        )
        assert res.ok, res.error
        assert seen == [expected], (name, seen)

        seen.clear()
        res = await spec_tool.execute(
            {"path": name, "sheet": "Q3", "spec": spec}, _ctx(ws)
        )
        assert res.ok, res.error
        assert seen == [expected], (name, seen)
