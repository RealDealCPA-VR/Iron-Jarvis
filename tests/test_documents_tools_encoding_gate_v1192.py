"""Document-tool defects from the 2026-08-20 sweep (findings 25 + 43).

25. ``convert_document`` csv->xlsx / csv->csv was the ONE csv read in the app
    still hard-coding ``encoding="utf-8", errors="replace"``. Every other csv
    read routes through ``readers._decode_bytes`` (utf-8-sig -> cp1252 ->
    charset-normalizer -> latin-1) precisely because Excel/Windows exports are
    usually BOM or cp1252 — so a legacy client list converted to .xlsx landed
    in the workbook as U+FFFD, while converting the SAME file to .docx (which
    goes through ``extract_text``) rendered it correctly.

43. ``extract_pdf`` sniffed the file's magic bytes BEFORE the ``fs_read_ok``
    gate, so a policy-denied OOXML-suffixed path was actually opened and read
    pre-deny, and which error came back ("not a PDF file" vs the denial
    reason) leaked one bit about its contents. Permissions fail CLOSED: the
    gate decides before anything touches the file.
"""

from __future__ import annotations

from pathlib import Path

from iron_jarvis.documents import document_tools
from iron_jarvis.tools.base import ToolContext


def _ctx(workspace: Path) -> ToolContext:
    return ToolContext(
        workspace=workspace,
        session_id="t",
        agent_run_id="t",
        config=None,
        event_bus=None,
        engine=None,
    )


def _tool(name: str):
    return next(t for t in document_tools() if t.name == name)


def _cells(xlsx: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        return [
            ["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()


# --- 25. legacy-CSV encodings survive the tabular conversion lane -------------


async def test_cp1252_csv_to_xlsx_keeps_its_accents(tmp_path):
    """A legacy Excel export is cp1252, not utf-8. The accented bytes must
    reach the workbook intact — under the old hard-coded utf-8 read they became
    U+FFFD replacement characters, silently, in the deliverable."""
    ws = tmp_path / "ws"
    ws.mkdir()
    src = ws / "clients.csv"
    src.write_bytes("name,city\nJosé García,Málaga\n".encode("cp1252"))

    res = await _tool("convert_document").execute(
        {"source": "clients.csv", "target": "clients.xlsx"}, _ctx(ws)
    )
    assert res.ok, res.error

    rows = _cells(ws / "clients.xlsx")
    assert rows[0] == ["name", "city"]
    assert rows[1] == ["José García", "Málaga"]
    assert not any("�" in c for row in rows for c in row)


async def test_utf8_bom_csv_to_xlsx_does_not_glue_the_bom_to_header(tmp_path):
    """Excel prepends a UTF-8 BOM. Read as plain utf-8 it survives as U+FEFF on
    the first header cell, so the column is named "﻿name" and every later
    lookup by header misses it."""
    ws = tmp_path / "ws"
    ws.mkdir()
    (ws / "bom.csv").write_bytes("name,amount\nRenée,12\n".encode("utf-8-sig"))

    res = await _tool("convert_document").execute(
        {"source": "bom.csv", "target": "bom.xlsx"}, _ctx(ws)
    )
    assert res.ok, res.error

    rows = _cells(ws / "bom.xlsx")
    assert rows[0][0] == "name"  # NOT "﻿name"
    assert rows[1][0] == "Renée"


async def test_cp1252_csv_to_csv_keeps_its_accents(tmp_path):
    """csv->csv rides the same tabular branch and had the same corruption."""
    ws = tmp_path / "ws"
    ws.mkdir()
    (ws / "in.csv").write_bytes("name\nMüller\n".encode("cp1252"))

    res = await _tool("convert_document").execute(
        {"source": "in.csv", "target": "out.csv"}, _ctx(ws)
    )
    assert res.ok, res.error

    out = (ws / "out.csv").read_bytes().decode("utf-8-sig")
    assert "Müller" in out
    assert "�" not in out


# --- 43. extract_pdf gates BEFORE it sniffs ----------------------------------


async def test_extract_pdf_denies_before_sniffing_the_file(tmp_path, monkeypatch):
    """A denied .xlsx must come back with the POLICY reason. Sniffing first
    answered "not a PDF file" for a denied non-PDF and the denial reason for a
    denied renamed PDF — a content oracle on a file the policy forbids."""
    allowed = tmp_path / "allowed"
    allowed.mkdir()
    denied = tmp_path / "outside" / "payroll.xlsx"
    denied.parent.mkdir(parents=True)
    denied.write_bytes(b"PK\x03\x04not-a-pdf-at-all")
    monkeypatch.setenv("IRONJARVIS_FS_ALLOWLIST", str(allowed))

    wsdir = tmp_path / "ws"
    wsdir.mkdir()
    res = await _tool("extract_pdf").execute({"path": str(denied)}, _ctx(wsdir))

    assert res.ok is False
    assert "IRONJARVIS_FS_ALLOWLIST" in res.error
    assert "not a PDF file" not in res.error


async def test_extract_pdf_does_not_read_bytes_of_a_denied_file(
    tmp_path, monkeypatch
):
    """The gate decides before anything opens the file. Recorded at the sniff
    itself: is_pdf_file must never be reached for a denied path."""
    from iron_jarvis.documents import ocr as ocr_mod

    seen: list[Path] = []
    real = ocr_mod.is_pdf_file

    def _spy(p):
        seen.append(Path(p))
        return real(p)

    monkeypatch.setattr(ocr_mod, "is_pdf_file", _spy)

    allowed = tmp_path / "allowed"
    allowed.mkdir()
    denied = tmp_path / "outside" / "secret.xlsx"
    denied.parent.mkdir(parents=True)
    denied.write_bytes(b"%PDF-1.7 renamed")  # the oracle case: really a PDF
    monkeypatch.setenv("IRONJARVIS_FS_ALLOWLIST", str(allowed))

    wsdir = tmp_path / "ws"
    wsdir.mkdir()
    res = await _tool("extract_pdf").execute({"path": str(denied)}, _ctx(wsdir))

    assert res.ok is False
    assert seen == []  # not sniffed at all — denied before any file access

    # ...and an ALLOWED path still reaches the content sniff (the v1.174.0
    # behaviour this must not cost us).
    good = allowed / "notes.txt"
    good.write_text("hello", encoding="utf-8")
    res2 = await _tool("extract_pdf").execute({"path": str(good)}, _ctx(wsdir))
    assert res2.ok is False and "not a PDF file" in res2.error
    assert seen == [good]
